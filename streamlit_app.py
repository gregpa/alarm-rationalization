"""
Alarm Rationalization Platform - Web Interface
Professional web application for transforming alarm data between DynAMo and PHA-Pro formats.
"""

import streamlit as st
import pandas as pd
import io
import csv
import os
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Set

# =============================================================================
# STRUCTURED LOGGING
# =============================================================================

class SessionLogHandler(logging.Handler):
    """Custom log handler that stores logs in Streamlit session state."""
    def emit(self, record):
        try:
            if 'app_logs' not in st.session_state:
                st.session_state.app_logs = []
            st.session_state.app_logs.append({
                'time': datetime.now().strftime('%H:%M:%S'),
                'level': record.levelname,
                'message': record.getMessage()
            })
            if len(st.session_state.app_logs) > 500:
                st.session_state.app_logs = st.session_state.app_logs[-500:]
        except Exception:
            pass

def setup_logger():
    logger = logging.getLogger('alarm_rationalization')
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        logger.addHandler(SessionLogHandler())
    return logger

app_logger = setup_logger()

# =============================================================================
# TRANSFORMATION HISTORY
# =============================================================================

def add_to_history(filename: str, direction: str, client: str, stats: Dict, output_data: bytes, output_filename: str):
    """Add a transformation to session history."""
    if 'transformation_history' not in st.session_state:
        st.session_state.transformation_history = []
    st.session_state.transformation_history.append({
        'id': len(st.session_state.transformation_history) + 1,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'input_file': filename,
        'direction': direction,
        'client': client,
        'tags': stats.get('tags', 0),
        'alarms': stats.get('alarms', 0),
        'output_data': output_data,
        'output_filename': output_filename
    })
    if len(st.session_state.transformation_history) > 20:
        st.session_state.transformation_history = st.session_state.transformation_history[-20:]

def get_history():
    if 'transformation_history' not in st.session_state:
        st.session_state.transformation_history = []
    return st.session_state.transformation_history

def clear_history():
    st.session_state.transformation_history = []

def clear_logs():
    st.session_state.app_logs = []

# =============================================================================
# EXPORT FORMAT HELPERS
# =============================================================================

def csv_to_excel(csv_data: bytes | str) -> bytes:
    """Convert CSV data to Excel format."""
    try:
        if isinstance(csv_data, bytes):
            df = pd.read_csv(io.BytesIO(csv_data), encoding='latin-1')
        else:
            df = pd.read_csv(io.StringIO(csv_data))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        app_logger.error(f"Excel conversion failed: {e}")
        raise

# =============================================================================
# EXTERNAL CONFIG LOADER (with fallback to hardcoded defaults)
# =============================================================================

def load_client_configs() -> Dict[str, Any]:
    """
    Load client configurations from external YAML file.
    Falls back to hardcoded defaults if YAML loading fails.

    Returns:
        Dict of client configurations
    """
    try:
        import yaml

        # Try to load from config/clients.yaml
        config_paths = [
            os.path.join(os.path.dirname(__file__), 'config', 'clients.yaml'),
            'config/clients.yaml',
            '/workspaces/alarm-rationalization/config/clients.yaml',
        ]

        for config_path in config_paths:
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    configs = yaml.safe_load(f)
                    if configs and isinstance(configs, dict):
                        # Successfully loaded from YAML
                        return configs

        # No config file found, fall back to hardcoded
        return None

    except Exception as e:
        # Any error loading YAML, fall back to hardcoded
        # This ensures the app always works even if YAML is corrupted
        return None


# Global config cache - loaded once at startup
_EXTERNAL_CONFIGS = load_client_configs()


def validate_client_configs(configs: Dict[str, Any]) -> List[Dict[str, str]]:
    """
    Validate client configurations and return list of warnings/errors.

    Returns:
        List of dicts with 'level' (warning/error), 'client', and 'message'
    """
    issues = []

    if not configs:
        return issues

    # Valid values for validation
    valid_parsers = {'dynamo', 'abb'}
    valid_unit_methods = {'TAG_PREFIX', 'ASSET_PARENT', 'ASSET_CHILD', 'FIXED'}
    valid_rule_types = {'exact', 'prefix', 'contains', 'in'}
    required_fields = ['name', 'parser', 'default_source']

    for client_id, config in configs.items():
        if not isinstance(config, dict):
            issues.append({
                'level': 'error',
                'client': client_id,
                'message': 'Configuration is not a valid dictionary'
            })
            continue

        # Check required fields
        for field in required_fields:
            if field not in config:
                issues.append({
                    'level': 'error',
                    'client': client_id,
                    'message': f'Missing required field: {field}'
                })

        # Validate parser type
        parser = config.get('parser', '')
        if parser and parser not in valid_parsers:
            issues.append({
                'level': 'error',
                'client': client_id,
                'message': f'Invalid parser "{parser}". Must be: {", ".join(valid_parsers)}'
            })

        # Validate unit_method
        unit_method = config.get('unit_method', '')
        if unit_method and unit_method not in valid_unit_methods:
            issues.append({
                'level': 'warning',
                'client': client_id,
                'message': f'Unknown unit_method "{unit_method}". Expected: {", ".join(valid_unit_methods)}'
            })

        # Validate TAG_PREFIX has unit_digits
        if unit_method == 'TAG_PREFIX' and 'unit_digits' not in config:
            issues.append({
                'level': 'warning',
                'client': client_id,
                'message': 'TAG_PREFIX method requires unit_digits (defaulting to 2)'
            })

        # Validate FIXED has unit_value
        if unit_method == 'FIXED' and 'unit_value' not in config:
            issues.append({
                'level': 'error',
                'client': client_id,
                'message': 'FIXED unit_method requires unit_value'
            })

        # Validate tag_source_rules
        rules = config.get('tag_source_rules', [])
        if rules:
            for i, rule in enumerate(rules):
                if not isinstance(rule, dict):
                    issues.append({
                        'level': 'error',
                        'client': client_id,
                        'message': f'tag_source_rules[{i}] is not a valid dictionary'
                    })
                    continue

                # Check rule has at least one match type
                match_types = set(rule.keys()) & valid_rule_types
                if not match_types:
                    issues.append({
                        'level': 'warning',
                        'client': client_id,
                        'message': f'tag_source_rules[{i}] has no match type ({", ".join(valid_rule_types)})'
                    })

                # Check rule has source
                if 'source' not in rule:
                    issues.append({
                        'level': 'error',
                        'client': client_id,
                        'message': f'tag_source_rules[{i}] missing "source" field'
                    })

        # Validate areas
        areas = config.get('areas', {})
        if areas:
            for area_id, area_config in areas.items():
                if not isinstance(area_config, dict):
                    issues.append({
                        'level': 'warning',
                        'client': client_id,
                        'message': f'Area "{area_id}" is not a valid dictionary'
                    })
                elif 'name' not in area_config:
                    issues.append({
                        'level': 'warning',
                        'client': client_id,
                        'message': f'Area "{area_id}" missing "name" field'
                    })

    return issues

# Global config validation cache
_CONFIG_WARNINGS = validate_client_configs(_EXTERNAL_CONFIGS) if _EXTERNAL_CONFIGS else []


def _preview_file_data(
    file_content: str,
    transformer: 'AlarmTransformer',
    direction: str,
    parser_type: str,
    selected_modes: List[str] = None
) -> Dict[str, Any]:
    """
    Preview file data without performing full transformation.
    Used by the "Preview before transform" feature.

    Args:
        file_content: The uploaded file content as string
        transformer: AlarmTransformer instance
        direction: "forward" or "reverse"
        parser_type: "dynamo" or "abb"

    Returns:
        Dict with preview statistics and potential issues
    """
    stats: Dict[str, Any] = {
        'total_rows': 0,
        'rows_to_process': 0,
        'rows_to_skip': 0,
        'units_found': set(),
        'issues': [],
        'skip_reasons': {}
    }

    try:
        lines = file_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        stats['total_rows'] = len([l for l in lines if l.strip()])

        if direction == "forward" and parser_type == "dynamo":
            # Analyze DynAMo file for forward transform
            empty_mode_valid = transformer.config.get("empty_mode_is_valid", False)

            for line in lines:
                if not line.strip():
                    continue

                # Check for _Parameter rows
                if "_Variable" in line and "_Parameter" in line:
                    # Parse the row to check mode
                    try:
                        row = list(csv.reader([line]))[0]
                        if len(row) >= 6:
                            mode = row[3].strip() if len(row) > 3 else ""
                            tag_name = row[1].strip() if len(row) > 1 else ""

                            # Check mode against selected modes or legacy behavior
                            mode_valid = False
                            if selected_modes is not None:
                                selected_upper = [m.upper() for m in selected_modes if m != "(empty)"]
                                allow_empty = "(empty)" in selected_modes
                                if mode.upper() in selected_upper or (allow_empty and mode == ""):
                                    mode_valid = True
                            else:
                                if mode.upper() == "NORMAL" or (empty_mode_valid and mode == ""):
                                    mode_valid = True

                            if mode_valid:
                                stats['rows_to_process'] += 1

                                # Extract unit for preview
                                if tag_name:
                                    unit = transformer.extract_unit(tag_name, "", "TAG_PREFIX")
                                    if unit:
                                        stats['units_found'].add(unit)
                            else:
                                stats['rows_to_skip'] += 1
                                reason = f"Mode: {mode}" if mode else "Empty mode"
                                stats['skip_reasons'][reason] = stats['skip_reasons'].get(reason, 0) + 1
                    except Exception:
                        pass

            # Check for potential issues
            if stats['rows_to_process'] == 0:
                stats['issues'].append("No NORMAL mode rows found - file may be empty or incorrectly formatted")

            if stats['rows_to_skip'] > stats['rows_to_process']:
                stats['issues'].append(f"More rows will be skipped ({stats['rows_to_skip']}) than processed ({stats['rows_to_process']})")

        elif direction == "reverse":
            # Analyze PHA-Pro file for reverse transform
            reader = csv.reader(lines)
            header = None

            for row in reader:
                if not row:
                    continue
                if header is None:
                    header = row
                    continue
                stats['rows_to_process'] += 1

            if stats['rows_to_process'] == 0:
                stats['issues'].append("No data rows found in PHA-Pro export")

        # Check for encoding issues
        if 'Ã' in file_content or 'â€' in file_content:
            stats['issues'].append("Possible encoding issues detected - special characters may not display correctly")

    except Exception as e:
        stats['issues'].append(f"Error analyzing file: {str(e)}")

    # Convert set to list for JSON serialization
    stats['units_found'] = list(stats['units_found'])

    return stats


# Page configuration - must be first Streamlit command
st.set_page_config(
    page_title="Alarm Rationalization Platform",
    page_icon="🔔",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# AUTHENTICATION
# =============================================================================

def check_password():
    """Returns True if the user has entered a valid password."""
    
    # Get authorized users from Streamlit Secrets
    # Format in secrets.toml:
    # [passwords]
    # user1 = "password1"
    # user2 = "password2"
    
    def validate_credentials(username, password):
        """Check if username/password combination is valid."""
        try:
            stored_passwords = st.secrets.get("passwords", {})
            if username in stored_passwords:
                return stored_passwords[username] == password
            return False
        except Exception:
            # If secrets not configured, allow access (for local development)
            st.warning("⚠️ Authentication not configured. Running in open mode.")
            return True
    
    # Check if already authenticated
    if st.session_state.get("authenticated", False):
        return True
    
    # Show login form
    st.markdown("""
    <style>
    .login-container {
        max-width: 400px;
        margin: 100px auto;
        padding: 40px;
        background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%);
        border-radius: 15px;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
    }
    .login-title {
        text-align: center;
        color: #ffffff;
        font-size: 1.8rem;
        margin-bottom: 10px;
    }
    .login-subtitle {
        text-align: center;
        color: #a0c4e8;
        font-size: 0.95rem;
        margin-bottom: 30px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("")
    st.markdown("")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown('<h1 class="login-title">🔔 Alarm Rationalization Platform</h1>', unsafe_allow_html=True)
        st.markdown('<p class="login-subtitle">Please log in to continue</p>', unsafe_allow_html=True)
        
        with st.form("login_form"):
            username = st.text_input("Username", placeholder="Enter your username")
            password = st.text_input("Password", type="password", placeholder="Enter your password")
            
            submitted = st.form_submit_button("Login", use_container_width=True)
            
            if submitted:
                if username and password:
                    if validate_credentials(username.lower().strip(), password):
                        st.session_state["authenticated"] = True
                        st.session_state["username"] = username.lower().strip()
                        st.rerun()
                    else:
                        st.error("âŒ Invalid username or password")
                else:
                    st.warning("Please enter both username and password")
        
        st.markdown("---")
        
        # Forgot password and Request Access links
        col_a, col_b = st.columns(2)
        
        with col_a:
            st.markdown(
                '<p style="text-align: center; color: #6c757d; font-size: 0.85rem;">'
                '📝‘ <a href="mailto:greg.pajak@aesolutions.com?subject=Password%20Reset%20Request%20-%20Alarm%20Platform">Forgot password?</a>'
                '</p>',
                unsafe_allow_html=True
            )
        
        with col_b:
            st.markdown(
                '<p style="text-align: center; color: #6c757d; font-size: 0.85rem;">'
                '📝 <a href="mailto:greg.pajak@aesolutions.com?subject=Access%20Request%20-%20Alarm%20Rationalization%20Platform&body=Hi%20Greg%2C%0A%0AI%20would%20like%20to%20request%20access%20to%20the%20Alarm%20Rationalization%20Platform.%0A%0AName%3A%20%0ACompany%3A%20%0AReason%20for%20access%3A%20%0A%0AThanks">Request access</a>'
                '</p>',
                unsafe_allow_html=True
            )
    
    return False

# Custom CSS for professional styling
st.markdown("""
<style>
    /* Main header styling */
    .main-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
    }
    .main-header h1 {
        margin: 0;
        font-size: 2.2rem;
        font-weight: 600;
    }
    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        font-size: 1.1rem;
    }
    
    /* Card styling */
    .card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        border: 1px solid #e0e0e0;
        margin-bottom: 1rem;
    }
    
    /* Status indicators */
    .status-success {
        background: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #28a745;
    }
    .status-info {
        background: #d1ecf1;
        color: #0c5460;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #17a2b8;
    }
    .status-warning {
        background: #fff3cd;
        color: #856404;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #ffc107;
    }
    
    /* Stats boxes */
    .stat-box {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 8px;
        text-align: center;
        border: 1px solid #e9ecef;
    }
    .stat-number {
        font-size: 2rem;
        font-weight: 700;
        color: #1e3a5f;
    }
    .stat-label {
        font-size: 0.9rem;
        color: #6c757d;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    /* Button styling */
    .stButton > button {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%);
        color: white;
        border: none;
        padding: 0.75rem 2rem;
        font-size: 1rem;
        font-weight: 500;
        border-radius: 8px;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(30, 58, 95, 0.3);
    }
    
    /* Sidebar styling */
    .css-1d391kg {
        background: #f8f9fa;
    }
    
    /* File uploader */
    .stFileUploader {
        border: 2px dashed #1e3a5f;
        border-radius: 10px;
        padding: 1rem;
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Professional table styling */
    .dataframe {
        font-size: 0.85rem;
    }

    /* Make selectbox dropdowns non-editable (selection only) */
    .stSelectbox input {
        caret-color: transparent !important;
        cursor: pointer !important;
    }
    .stSelectbox [data-baseweb="select"] input {
        caret-color: transparent !important;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# CORE TRANSFORMATION LOGIC (embedded for single-file deployment)
# ============================================================

class AlarmTransformer:
    """Core transformation engine."""
    
    # PHA-Pro 45-column headers
    PHAPRO_HEADERS = [
        "Unit", "Tag Name", "Old Tag Description", "New Tag Description", "P&ID",
        "Range Min", "Range Max", "Engineering Units", "Tag Source",
        "Rationalization (Tag) Comment", "Old Tag Enable Status", "New Tag Enable Status",
        "Alarm Type", "Old Individual Alarm Enable Status", "New Individual Alarm Enable Status",
        "Old (BPCS) Priority", "New (BPCS) Priority", "Old Limit", "New Limit",
        "Old Deadband", "New Deadband", "Old Deadband Units", "New Deadband Units",
        "Old On-Delay Time", "New On-Delay Time", "Old Off-Delay Time", "New Off-Delay Time",
        "Rationalization Status", "Alarm Status", "Rationalization (Alarm) Comment",
        "Limit Owner", "Alarm HAZOP Comment", "Alarm Suppression Notes", "Alarm Class",
        "Cause(s)", "Consequence(s)", "Inside Action(s)", "Outside Action(s)",
        "Health and Safety", "Environment", "Financial", "Reputation",
        "Privilege to Operate", "Max Severity", "Allowable Time to Respond"
    ]
    
    # DynAMo 42-column headers
    DYNAMO_HEADERS = [
        "'_Variable", "name", "_Parameter", "mode", "boundary", "alarmType", "alarmName",
        "value", "enforcement", "priorityName", "priorityValue", "priorityEnforcement",
        "consequence", "TimeToRespond", "pre-Alarm", "historyTag", "Purpose of Alarm",
        "Consequence of No Action", "Board Operator", "Field Operator", "Supporting Notes",
        "TypeParameter", "TypeValue", "TypeEnforcement", "DisabledParameter", "DisabledValue",
        "DisabledEnforcement", "SuppressedParameter", "SuppressedValue", "SuppressedEnforcement",
        "OnDelayParameter", "OnDelayValue", "OnDelayEnforcement", "OffDelayParameter",
        "OffDelayValue", "OffDelayEnforcement", "DeadBandParameter", "DeadBandValue",
        "DeadBandEnforcement", "DeadBandUnitParameter", "DeadBandUnitValue", "DeadBandUnitEnforcement"
    ]
    
    # ABB PHA-Pro 23-column headers (different from DynAMo's 45-column)
    ABB_PHAPRO_HEADERS = [
        "Unit", "Starting Tag Name", "New Tag Name", "Old Tag Description", "New Tag Description",
        "Tag Source", "Rationalization (Tag) Comment", "Range Min", "Range Max", "Engineering Units",
        "Starting Alarm Type", "New Alarm Type", "Old Alarm Enable Status", "New Alarm Enable Status",
        "Old Alarm Severity", "New Alarm Severity", "Old Limit", "New Limit",
        "Old (BPCS) Priority", "New (BPCS) Priority", "Rationalization Status", "Alarm Status",
        "Rationalization (Alarm) Comment"
    ]
    
    # ABB Return format (8 columns)
    ABB_RETURN_HEADERS = [
        "Tag Name", "New Tag Description", "New Alarm Type", "New Alarm Enable Status",
        "New Limit", "New Priority", "New Alarm Severity Level", "ABB Consolidated Notes"
    ]
    
    # HF Sinclair PHA-Pro 43-column headers (matches Tag_Import template)
    HFS_PHAPRO_HEADERS = [
        "Unit", "Starting Tag Name", "New Tag Name", "Old Tag Description", "New Tag Description",
        "P&ID", "Range Min", "Range Max", "Engineering Units", "Tag Source", "Rationalization (Tag) Comment",
        "Old Tag Enable Status", "New Tag Enable Status",
        "Starting Alarm Type", "New Alarm Type", "Old Alarm Enable Status", "New Alarm Enable Status",
        "Old (BPCS) Priority", "New (BPCS) Priority",
        "Old Limit", "New Limit", "Old Deadband", "New Deadband", "Old Deadband Units", "New Deadband Units",
        "Old On-Delay Time", "New On-Delay Time", "Old Off-Delay Time", "New Off-Delay Time",
        "Rationalization Status", "Alarm Status", "Rationalization (Alarm) Comment", "Alarm Class",
        "Cause(s)", "Consequence(s)", "Inside Action(s)", "Outside Action(s)",
        "Escalation", "Limit Owner", "Personnel", "Public or Environment", "Costs / Production",
        "Maximum Time to Resolve"
    ]
    
    # Client configurations with Unit/Area hierarchy
    # NOTE: These are FALLBACK defaults. External configs from config/clients.yaml take precedence.
    # Edit config/clients.yaml instead of modifying these values directly.
    _HARDCODED_CONFIGS = {
        "flng": {
            "name": "Freeport LNG",
            "vendor": "Honeywell Experion/DynAMo",
            "dcs_name": "DynAMo",
            "pha_tool": "PHA-Pro",
            "parser": "dynamo",
            "unit_method": "TAG_PREFIX",
            "unit_digits": 2,
            "tag_source_rules": [
                {"prefix": "SM", "field": "point_type", "source": "Honeywell Safety Manager (SIS)", "enforcement": "R"},
                {"contains": ".", "field": "tag_name", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"in": ["ANA", "STA"], "field": "point_type", "source": "Honeywell Experion (SCADA)", "enforcement": "M"},
            ],
            "default_source": "Honeywell TDC (DCS)",
            "areas": {
                "lqf_u17": {"name": "LQF - Unit 17", "description": "Liquefaction Facility Unit 17"},
                "ptf_u61": {"name": "PTF - Unit 61", "description": "Pretreatment Facility Unit 61"},
            },
            "default_area": "lqf_u17",
        },
        "hfs_artesia": {
            "name": "HF Sinclair - Artesia",
            "vendor": "Honeywell Experion/DynAMo",
            "dcs_name": "DynAMo",
            "pha_tool": "PHA-Pro",
            "parser": "dynamo",
            "unit_method": "TAG_PREFIX",
            "unit_digits": 2,
            # Tag source rules based on Point Type (from Honeywell_tag_info.xlsx reference)
            # Evaluated in order - first match wins
            "tag_source_rules": [
                # Safety Manager point types
                {"prefix": "SM_", "field": "point_type", "source": "Honeywell Safety Manager", "enforcement": "R"},
                {"exact": "SM", "field": "point_type", "source": "Honeywell Safety Manager", "enforcement": "R"},
                # SCADA point types
                {"exact": "ANA", "field": "point_type", "source": "Honeywell Experion (SCADA)", "enforcement": "M"},
                {"exact": "STA", "field": "point_type", "source": "Honeywell Experion (SCADA)", "enforcement": "M"},
                # Experion DCS point types
                {"exact": "AUTOMAN", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "CAB", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "DATAACQ", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "DEVCTL", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "DIGACQ", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "FANOUT", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "FLAG/CONTACTMON", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "FLOWCOMP", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "OVRDSEL", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "PID", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "PID-PL", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "PIDFF", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "RATIOCTL", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "REGCALC", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "REMCAS", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "SIGNALSEL", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "SWITCH", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                {"exact": "TOTALIZER", "field": "point_type", "source": "Honeywell Experion (DCS)", "enforcement": "M"},
                # TDC point types
                {"exact": "ANALGIN", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "CUSTOMAM", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "DIGCOM", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "DIGIN", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "FLAGAM", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "LOGIC", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "REGAM", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "REGCTL", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "REGLATRY", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "REGPV", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
                {"exact": "SWITCHAM", "field": "point_type", "source": "Honeywell TDC (DCS)", "enforcement": "M"},
            ],
            "default_source": "Honeywell TDC (DCS)",
            "empty_mode_is_valid": True,
            "phapro_headers": "HFS",
            "areas": {
                "north_console": {"name": "North Console", "description": "North Console Area"},
            },
            "default_area": "north_console",
        },
        "rt_bessemer": {
            "name": "Rio Tinto - Bessemer City",
            "vendor": "ABB",
            "dcs_name": "ABB",
            "pha_tool": "PHA-Pro",
            "parser": "abb",  # Use ABB parser
            "unit_method": "FIXED",
            "unit_value": "Line 1",  # Fixed unit for this site
            "tag_source_rules": [],
            "default_source": "ABB 800xA (DCS)",
            # ABB-specific alarm type mappings
            "abb_alarm_types": {
                "H": "(PV) High",
                "HH": "(PV) High High", 
                "HHH": "(PV) High High High",
                "L": "(PV) Low",
                "LL": "(PV) Low Low",
                "LLL": "(PV) Low Low Low",
                "OE": "Object Error",
            },
            "abb_priority_default": 3,
            "areas": {
                "line_1": {"name": "Line 1", "description": "Production Line 1"},
            },
            "default_area": "line_1",
        },
    }

    @classmethod
    def get_client_configs(cls) -> Dict[str, Any]:
        """
        Get client configurations - external YAML if available, otherwise hardcoded fallback.

        Returns:
            Dict of client configurations
        """
        global _EXTERNAL_CONFIGS
        if _EXTERNAL_CONFIGS is not None:
            return _EXTERNAL_CONFIGS
        return cls._HARDCODED_CONFIGS

    # Property to maintain backward compatibility with code that references CLIENT_CONFIGS
    @property
    def CLIENT_CONFIGS(self) -> Dict[str, Any]:
        """Get client configs (instance property for backward compatibility)."""
        return self.get_client_configs()

    # ABB-specific column mappings
    ABB_ALARM_SUFFIXES = ['H', 'HH', 'HHH', 'L', 'LL', 'LLL', 'OE']
    
    DISCRETE_ALARM_TYPES = [
        "controlfail", "st0", "st1", "st2", "st3", "unreasonable", "bad pv",
        "off normal", "command disagree", "command fail", "cnferr", "chofst", "offnrm",
        # Additional discrete patterns found in FLNG project
        "bad control", "override interlock", "safety interlock", "safety override",
        "uncommanded", "c1 -", "c2 -", "c3 -", "c4 -", "c5 -", "c6 -", 
        "c7 -", "c8 -", "c9 -", "c10 -", "c11 -", "c12 -",
        "flagoffnorm", "devbadpv", "devcmddis", "devuncevt", "devcmdfail",
        "daqpvhi", "daqpvhihi", "daqpvlow", "daqpvlolo", "daqrocneg", "daqrocpos", "regbadctl"
    ]
    
    def __init__(self, client_id: str, area_id: str = None):
        """Initialize transformer with client and optional area configuration."""
        self.client_id = client_id
        base_config = self.CLIENT_CONFIGS.get(client_id, self.CLIENT_CONFIGS["flng"])
        
        # Start with base config (copy to avoid mutation)
        self.config = dict(base_config)
        
        # Determine area
        self.area_id = area_id or base_config.get("default_area")
        
        # If area specified, merge area-specific overrides
        areas = base_config.get("areas", {})
        if self.area_id and self.area_id in areas:
            area_config = areas[self.area_id]
            for key, value in area_config.items():
                if key not in ["name", "description"]:
                    self.config[key] = value
            self.area_name = area_config.get("name", self.area_id)
        else:
            self.area_name = None
        
        self.stats = {"tags": 0, "alarms": 0, "units": set()}
    
    @classmethod
    def get_client_areas(cls, client_id: str) -> dict:
        """Get available areas for a client."""
        client_config = cls.get_client_configs().get(client_id, {})
        areas = client_config.get("areas", {})
        return {aid: aconfig.get("name", aid) for aid, aconfig in areas.items()}
    
    def get_phapro_headers(self) -> list:
        """Get the PHA-Pro headers for this client format."""
        phapro_format = self.config.get("phapro_headers", "FLNG")
        if phapro_format == "HFS":
            return self.HFS_PHAPRO_HEADERS
        else:
            return self.PHAPRO_HEADERS
    
    def parse_abb_excel(self, file_bytes: bytes) -> List[Dict]:
        """Parse ABB 800xA Excel export (wide format with alarm columns)."""
        import pandas as pd
        
        df = pd.read_excel(io.BytesIO(file_bytes))
        
        tags = []
        
        # Column name mappings (handle variations)
        name_col = None
        desc_col = None
        for col in df.columns:
            col_upper = col.upper()
            if col_upper == 'NAME' or col_upper == 'OBJECT NAME':
                name_col = col
            if col_upper == 'DESCRIPTION':
                desc_col = col
        
        if not name_col:
            name_col = df.columns[3] if len(df.columns) > 3 else df.columns[0]
        if not desc_col:
            desc_col = df.columns[4] if len(df.columns) > 4 else None
        
        for _, row in df.iterrows():
            tag_name = str(row.get(name_col, '')).strip()
            if not tag_name or tag_name == 'nan':
                continue
            
            tag_data = {
                'tag_name': tag_name,
                'description': str(row.get(desc_col, '')).strip() if desc_col else '',
                'alarms': []
            }
            
            # Extract alarm data for each type (H, HH, HHH, L, LL, LLL, OE)
            alarm_types = self.config.get('abb_alarm_types', {
                'H': '(PV) High', 'HH': '(PV) High High', 'HHH': '(PV) High High High',
                'L': '(PV) Low', 'LL': '(PV) Low Low', 'LLL': '(PV) Low Low Low',
                'OE': 'Object Error'
            })
            
            for suffix, alarm_name in alarm_types.items():
                # Find columns for this alarm type
                conf_col = None
                level_col = None
                sev_col = None
                
                for col in df.columns:
                    col_upper = col.upper()
                    if f'AECONF{suffix}' in col_upper or f'AECONF{suffix.upper()}' == col_upper:
                        conf_col = col
                    if f'AELEVEL{suffix}' in col_upper or f'AELEVEL{suffix.upper()}' == col_upper:
                        level_col = col
                    if f'AESEV{suffix}' in col_upper or f'AESEV{suffix.upper()}' == col_upper:
                        sev_col = col
                
                # Get values
                enabled = 0
                level = -9999999
                severity = 1
                
                if conf_col:
                    try:
                        enabled = int(row.get(conf_col, 0))
                    except:
                        enabled = 0
                
                if level_col:
                    try:
                        level = float(row.get(level_col, -9999999))
                    except:
                        level = -9999999
                
                if sev_col:
                    try:
                        severity = int(row.get(sev_col, 1))
                    except:
                        severity = 1
                
                tag_data['alarms'].append({
                    'type': alarm_name,
                    'suffix': suffix,
                    'enabled': enabled,
                    'level': level,
                    'severity': severity,
                    'priority': self.config.get('abb_priority_default', 3),
                })
            
            tags.append(tag_data)
        
        return tags
    
    def parse_dynamo_csv(self, file_content: str) -> Dict:
        """Parse DynAMo multi-schema CSV file."""
        lines = file_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        reader = csv.reader(lines)
        
        schemas = {
            '_DCSVariable': {},
            '_DCS': {},
            '_Parameter': {},
            '_Notes': {},
        }
        
        for row in reader:
            if not row or len(row) < 3:
                continue
            
            if row[0].strip() == "_Variable":
                tag_name = row[1].strip()
                schema_type = row[2].strip()
                
                if schema_type == "_DCSVariable":
                    schemas['_DCSVariable'][tag_name] = {
                        'assetPath': row[3] if len(row) > 3 else "",
                        'pointType': row[8] if len(row) > 8 else "",
                    }
                elif schema_type == "_DCS":
                    schemas['_DCS'][tag_name] = {
                        'engUnits': row[3] if len(row) > 3 else "",
                        'pointType': row[4] if len(row) > 4 else "",
                        'PVEUHI': row[5] if len(row) > 5 else "",
                        'PVEULO': row[6] if len(row) > 6 else "",
                        'desc': row[7] if len(row) > 7 else "",
                        'unit': row[10] if len(row) > 10 else "",  # Full unit name from column 10
                    }
                elif schema_type == "_Parameter":
                    # Only store NORMAL mode parameters for forward transform
                    mode = row[3] if len(row) > 3 else "Base"
                    if tag_name not in schemas['_Parameter']:
                        schemas['_Parameter'][tag_name] = []
                    schemas['_Parameter'][tag_name].append({
                        'mode': mode,
                        'alarmType': row[5] if len(row) > 5 else "",
                        'value': row[7] if len(row) > 7 else "",
                        'priorityValue': row[10] if len(row) > 10 else "",
                        'consequence': row[12] if len(row) > 12 else "",
                        'TimeToRespond': row[13] if len(row) > 13 else "",
                        'PurposeOfAlarm': row[16] if len(row) > 16 else "",
                        'ConsequenceOfNoAction': row[17] if len(row) > 17 else "",
                        'BoardOperator': row[18] if len(row) > 18 else "",
                        'FieldOperator': row[19] if len(row) > 19 else "",
                        'SupportingNotes': row[20] if len(row) > 20 else "",
                        'DisabledValue': row[25] if len(row) > 25 else "",
                        'OnDelayValue': row[31] if len(row) > 31 else "",
                        'OffDelayValue': row[34] if len(row) > 34 else "",
                        'DeadBandValue': row[37] if len(row) > 37 else "",
                        'DeadBandUnitValue': row[40] if len(row) > 40 else "",
                    })
                elif schema_type == "_Notes":
                    schemas['_Notes'][tag_name] = {
                        'DocRef1': row[11] if len(row) > 11 else "",
                    }
        
        return schemas
    
    def extract_unit(self, tag_name: str, asset_path: str = "", method: str = None) -> str:
        """Extract unit from tag name or asset path.
        
        Args:
            tag_name: The tag name
            asset_path: The asset path (optional)
            method: Override method - "tag_prefix", "asset_parent", "asset_child" (optional)
        
        Returns:
            Unit string. Returns "00" if no unit found (for TAG_PREFIX method).
        
        Methods:
            - tag_prefix: First digits of tag name (e.g., "17" from "17TI5879")
            - asset_parent: First level after /U##/ (e.g., "17_FLARE" from /U17/17_FLARE/17H-2)
            - asset_child: Last level in path (e.g., "17H-2" from /U17/17_FLARE/17H-2)
        """
        import re
        
        # Determine which method to use
        use_method = method or self.config.get("unit_method", "TAG_PREFIX")
        use_method = use_method.upper()
        
        # Extract unit from tag prefix
        unit_from_prefix = ""
        for ch in tag_name:
            if ch.isdigit():
                unit_from_prefix += ch
                if len(unit_from_prefix) >= self.config["unit_digits"]:
                    break
            elif unit_from_prefix:
                break
        
        # If no digits found at start of tag, use "00" as default unit
        if not unit_from_prefix:
            unit_from_prefix = "00"
        
        # Extract parent and child units from asset path
        unit_parent = ""
        unit_child = ""
        if asset_path:
            # Parse asset path: /Assets/LQF/U17/17_FLARE/17H-2
            match = re.search(r'/U(\d+)/', asset_path, re.IGNORECASE)
            if match:
                # Get everything after /U##/
                u_pos = match.end()
                remaining = asset_path[u_pos:]
                
                # Split by /
                parts = [p for p in remaining.split('/') if p]
                
                if len(parts) >= 1:
                    # Parent unit is first level after U##
                    unit_parent = parts[0]
                
                if len(parts) >= 2:
                    # Child unit is last level
                    unit_child = parts[-1]
                elif len(parts) == 1:
                    # No child level, parent is also the "child"
                    unit_child = parts[0]
        
        # Return based on method
        if use_method == "TAG_PREFIX":
            return unit_from_prefix
        elif use_method == "ASSET_PARENT":
            return unit_parent if unit_parent else "00"
        elif use_method == "ASSET_CHILD":
            return unit_child if unit_child else (unit_parent if unit_parent else "00")
        elif use_method == "ASSET_PATH":
            # Legacy - use parent
            return unit_parent if unit_parent else "00"
        elif use_method == "BOTH":
            # Both tag prefix and asset parent must match
            if unit_from_prefix and unit_parent and unit_from_prefix in unit_parent:
                return unit_parent
            return "00"
        
        return unit_from_prefix  # default fallback
    
    def derive_tag_source(self, tag_name: str, point_type: str) -> Tuple[str, str]:
        """Derive tag source and enforcement from rules."""
        pt_upper = point_type.upper() if point_type else ""
        
        for rule in self.config.get("tag_source_rules", []):
            field = rule.get("field", "point_type")
            check_value = pt_upper if field == "point_type" else tag_name
            
            # Exact match
            if "exact" in rule:
                if check_value == rule["exact"].upper():
                    return rule["source"], rule.get("enforcement", "M")
            # Prefix match
            elif "prefix" in rule:
                if check_value.startswith(rule["prefix"].upper()):
                    return rule["source"], rule.get("enforcement", "M")
            # Contains match
            elif "contains" in rule:
                if rule["contains"] in check_value:
                    return rule["source"], rule.get("enforcement", "M")
            # In list match
            elif "in" in rule:
                if check_value in [v.upper() for v in rule["in"]]:
                    return rule["source"], rule.get("enforcement", "M")
        
        return self.config.get("default_source", "Unknown"), "M"
    
    def map_priority(self, priority: str, disabled_value: str = "") -> Tuple[str, str]:
        """Map priority to code and alarm status."""
        p = priority.strip().lower() if priority else ""
        
        mapping = {
            'urgent': ('U', 'Alarm'),
            'critical': ('C', 'Alarm'),
            'high': ('H', 'Alarm'),
            'medium': ('M', 'Alarm'),
            'low': ('L', 'Alarm'),
            'journal': ('J', 'Event'),
            'none': ('N', 'None'),
            'noaction': ('NA', 'None'),  # HFS specific
        }
        
        code, status = mapping.get(p, ('N', 'None'))
        
        # Jo for disabled Journal alarms
        if code == 'J' and disabled_value.upper() == 'FALSE':
            code = 'Jo'
        
        return code, status
    
    def map_severity(self, consequence: str) -> str:
        """Map consequence text to severity code (A-E or (N))."""
        if not consequence or consequence.strip() in ["~", "", "-"]:
            return "(N)"
        
        c = consequence.strip().upper()
        
        # Already a letter code
        if c in ['A', 'B', 'C', 'D', 'E']:
            return c
        
        # Text to letter mapping
        text_mapping = {
            'CATASTROPHIC': 'A',
            'MAJOR': 'B', 
            'MODERATE': 'C',
            'MINOR': 'D',
            'INSIGNIFICANT': 'E',
        }
        
        # Check for text matches (including partial)
        for text, code in text_mapping.items():
            if text in c or c in text:
                return code
        
        # No match found
        return "(N)"
    
    def is_discrete(self, alarm_type: str) -> bool:
        """Check if alarm type is discrete."""
        at_lower = alarm_type.lower()
        return any(d in at_lower for d in self.DISCRETE_ALARM_TYPES)
    
    def _clean_value(self, value: str) -> str:
        """Clean a value - return empty string for placeholder values like ~.
        Also removes commas from numeric values and fixes encoding issues.
        """
        if not value or value.strip() in ['~', '-', '']:
            return ""
        cleaned = value.strip()
        # Remove commas from numeric values (e.g., "1,500" -> "1500")
        cleaned = cleaned.replace(',', '')
        return self._fix_encoding(cleaned)
    
    def _fix_encoding(self, value: str) -> str:
        """Fix common encoding issues, particularly the degree symbol.
        
        When Latin-1 encoded files are read and then written to UTF-8,
        characters like ° can get double-encoded (Ã‚°).
        This fixes those issues.
        """
        if not value:
            return value
        
        # Common encoding fixes
        # These occur when UTF-8 bytes are interpreted as Latin-1
        replacements = {
            'Ã‚°': '°',      # Double-encoded degree symbol (UTF-8 0xC2 0xB0 as Latin-1)
            'Ã‚\xa0': '\xa0',  # Non-breaking space (UTF-8 0xC2 0xA0 as Latin-1)
            'Ã‚ ': ' ',      # Another form of space encoding issue
            'Ã¢â‚¬â„¢': "'",     # Smart quote
            'Ã¢â‚¬"': '—',     # En dash
            'Ã¢â‚¬"': '–',     # Em dash
            'ÃƒÂ©': 'Ã©',      # Accented e
            'ÃƒÂ±': 'Ã±',      # Spanish n
        }
        
        result = value
        for bad, good in replacements.items():
            result = result.replace(bad, good)
        
        return result
    
    def validate_phapro_columns(self, col_map: Dict[str, int]) -> List[str]:
        """
        Validate that all required PHA-Pro columns are present.
        Supports flexible column name mapping for different clients.
        
        Returns:
            List of missing column names (empty if all present)
        """
        phapro_format = self.config.get("phapro_headers", "FLNG")
        
        if phapro_format == "HFS":
            # HFS PHA-Pro export has different column names
            # More flexible - check for alternate names
            required_columns = {
                'Tag Name': ['Tag Name', 'New Tag Name', 'Starting Tag Name'],
                'Alarm Type': ['Alarm Type', 'New Alarm Type', 'Starting Alarm Type'],
                'New Limit': ['New Limit'],
                'New Priority': ['New Priority', 'New (BPCS) Priority'],
                'Cause(s)': ['Cause(s)'],
                'Consequence(s)': ['Consequence(s)'],
                'Inside Action(s)': ['Inside Action(s)'],
                'Outside Action(s)': ['Outside Action(s)'],
                'Max Severity': ['Max Severity'],
                'TTR Range': ['TTR Range', 'Max Time to Resolve', 'Allowable Time to Respond'],
            }
            # Rationalization Status is optional but useful - can derive alarm status
        else:
            # FLNG format - stricter column names
            required_columns = {
                'Tag Name': ['Tag Name'],
                'Tag Source': ['Tag Source'],
                'Alarm Type': ['Alarm Type'],
                'New Priority': ['New Priority', 'New (BPCS) Priority'],
                'New Limit': ['New Limit'],
                'Alarm Status': ['Alarm Status'],
                'Cause(s)': ['Cause(s)'],
                'Consequence(s)': ['Consequence(s)'],
                'Inside Action(s)': ['Inside Action(s)'],
                'Outside Action(s)': ['Outside Action(s)'],
                'Max Severity': ['Max Severity'],
                'TTR Range': ['TTR Range', 'Allowable Time to Respond'],
                'New Individual Alarm Enable Status': ['New Individual Alarm Enable Status'],
            }
        
        # Check for missing columns - a column is present if ANY of its alternates exist
        missing = []
        for col_name, alternates in required_columns.items():
            found = False
            for alt in alternates:
                if alt in col_map:
                    found = True
                    break
            if not found:
                missing.append(col_name)
        
        return missing
    
    def _get_col_flexible(self, col_map: Dict[str, int], row: list, names: list, default: str = "") -> str:
        """Get column value trying multiple possible column names."""
        for name in names:
            idx = col_map.get(name)
            if idx is not None and idx < len(row):
                val = row[idx].strip()
                if val:
                    return val
        return default
    
    def get_required_columns_info(self) -> Dict[str, str]:
        """Return dictionary of required columns and their purposes."""
        phapro_format = self.config.get("phapro_headers", "FLNG")
        
        if phapro_format == "HFS":
            return {
                'Tag Name (or New Tag Name)': 'Tag identifier - needed to map back to DynAMo',
                'Alarm Type (or New Alarm Type)': 'Required to identify which alarm parameter to update',
                'New Priority': 'Maps to DynAMo priorityValue',
                'New Limit': 'Maps to DynAMo value field for analog alarms',
                'Cause(s)': 'Maps to DynAMo Purpose of Alarm',
                'Consequence(s)': 'Maps to DynAMo Consequence of No Action',
                'Inside Action(s)': 'Maps to DynAMo Board Operator',
                'Outside Action(s)': 'Maps to DynAMo Field Operator',
                'Max Severity': 'Maps to DynAMo consequence field',
                'TTR Range': 'Maps to DynAMo TimeToRespond',
            }
        else:
            return {
                'Tag Name': 'Tag identifier - needed to map back to DynAMo',
                'Tag Source': 'Determines enforcement (M vs R for Safety Manager)',
                'Alarm Type': 'Required to identify which alarm parameter to update',
                'New Priority': 'Maps to DynAMo priorityValue',
                'New Limit': 'Maps to DynAMo value field for analog alarms',
                'Alarm Status': 'Determines consequence and disabled state',
                'Cause(s)': 'Maps to DynAMo Purpose of Alarm',
                'Consequence(s)': 'Maps to DynAMo Consequence of No Action',
                'Inside Action(s)': 'Maps to DynAMo Board Operator',
                'Outside Action(s)': 'Maps to DynAMo Field Operator',
                'Max Severity': 'Maps to DynAMo consequence field',
                'TTR Range': 'Maps to DynAMo TimeToRespond',
                'New Individual Alarm Enable Status': 'Maps to DynAMo DisabledValue (TRUE/FALSE)',
            }
    
    def transform_forward(self, file_content: str, selected_units: List[str] = None, unit_method: str = None, selected_modes: List[str] = None) -> Tuple[str, Dict]:
        """Transform DynAMo to PHA-Pro format.

        Args:
            file_content: The CSV file content
            selected_units: List of units to filter (optional)
            unit_method: "tag_prefix" or "asset_path" (optional, uses config default)
        """
        app_logger.info(f"Forward transform started - client: {self.client_id}, units: {selected_units}")
        schemas = self.parse_dynamo_csv(file_content)
        
        rows = []
        self.stats = {"tags": 0, "alarms": 0, "units": set(), "skipped_modes": 0}
        
        # Build tag list sorted by unit and name
        tags = []
        for tag_name in schemas['_DCSVariable'].keys():
            var_data = schemas['_DCSVariable'].get(tag_name, {})
            dcs_data = schemas['_DCS'].get(tag_name, {})
            params = schemas['_Parameter'].get(tag_name, [])
            notes = schemas['_Notes'].get(tag_name, {})
            
            if not params:
                continue
            
            # Filter by selected modes (user-configurable)
            # If selected_modes provided, use those; otherwise fall back to config behavior
            if selected_modes is not None:
                # User explicitly selected modes - use their selection
                # Handle "(empty)" as a special token for blank mode values
                selected_upper = [m.upper() for m in selected_modes if m != "(empty)"]
                allow_empty = "(empty)" in selected_modes
                normal_params = [
                    p for p in params
                    if p.get('mode', '').upper() in selected_upper
                    or (allow_empty and p.get('mode', '').strip() == '')
                ]
            else:
                # No explicit selection - use legacy behavior
                empty_mode_is_valid = self.config.get("empty_mode_is_valid", False)
                if empty_mode_is_valid:
                    normal_params = [p for p in params if p.get('mode', '').upper() in ['NORMAL', '']]
                else:
                    normal_params = [p for p in params if p.get('mode', '').upper() == 'NORMAL']
            skipped = len(params) - len(normal_params)
            self.stats["skipped_modes"] += skipped
            
            if not normal_params:
                continue
            
            point_type = dcs_data.get('pointType', '') or var_data.get('pointType', '')
            asset_path = var_data.get('assetPath', '')
            
            # Extract unit based on selected method
            extracted_unit = self.extract_unit(tag_name, asset_path, unit_method)
            
            # For backward compatibility, also check _DCS[10] if method is asset_child
            dcs_unit = dcs_data.get('unit', '')
            
            # Determine final unit based on method
            if unit_method and unit_method.upper() == "ASSET_PARENT":
                # Use parent unit from asset path extraction
                final_unit = extracted_unit
            elif unit_method and unit_method.upper() == "ASSET_CHILD":
                # Prefer _DCS[10] if available (it's the most specific), otherwise use extraction
                final_unit = dcs_unit if dcs_unit else extracted_unit
            else:
                # Tag prefix or default
                final_unit = extracted_unit
            
            # For unit filtering with asset methods, we need to check if selected_units 
            # matches either the extracted unit or is contained in it
            if selected_units:
                match_found = False
                for sel_unit in selected_units:
                    if sel_unit == final_unit or sel_unit in final_unit or final_unit in sel_unit:
                        match_found = True
                        break
                if not match_found:
                    continue
            
            self.stats["units"].add(final_unit)
            
            # Get engineering units - prefer _DCS, fall back to _DCSVariable
            # Apply encoding fix for degree symbol and other characters
            # Leave blank if ~ or empty (don't use placeholder values)
            eng_units = dcs_data.get('engUnits', '') or var_data.get('engUnits', '')
            eng_units = self._fix_encoding(eng_units)
            if eng_units in ['~', '-', '']:
                eng_units = ''
            
            # Clean range values - leave blank if ~ or empty or -------- (don't use default 0/1)
            range_min = dcs_data.get('PVEULO', '').replace(',', '')
            range_max = dcs_data.get('PVEUHI', '').replace(',', '')
            if range_min in ['~', '-', '--------', '']:
                range_min = ''
            if range_max in ['~', '-', '--------', '']:
                range_max = ''
            
            # P&ID - use "UNKNOWN" if not available
            pid = notes.get('DocRef1', '')
            if not pid or pid in ['~', '']:
                pid = 'UNKNOWN'
            
            tags.append({
                'tag_name': tag_name,
                'unit': final_unit,
                'point_type': point_type,
                'desc': self._fix_encoding(dcs_data.get('desc', '')),
                'eng_units': eng_units,
                'range_min': range_min,
                'range_max': range_max,
                'pid': pid,
                'params': normal_params,  # Only NORMAL mode params
            })
        
        # Sort by unit, then tag name
        tags.sort(key=lambda t: (t['unit'], t['tag_name']))
        
        last_unit = None
        
        for tag in tags:
            self.stats["tags"] += 1
            tag_source, enforcement = self.derive_tag_source(tag['tag_name'], tag['point_type'])
            
            is_first_tag_for_unit = (tag['unit'] != last_unit)
            is_first_alarm_for_tag = True
            
            for param in tag['params']:
                alarm_type = param.get('alarmType', '').strip()
                # Skip rows with no alarm type or placeholder values
                if not alarm_type or alarm_type in ['~', '-', '']:
                    continue
                
                self.stats["alarms"] += 1
                priority_code, alarm_status = self.map_priority(
                    param.get('priorityValue', ''),
                    param.get('DisabledValue', '')
                )
                
                # Derive individual alarm enable status
                at_lower = param.get('alarmType', '').lower()
                disabled_val = param.get('DisabledValue', '').upper()
                
                # ControlFail and certain discrete alarms use {n/a}
                if self.is_discrete(param.get('alarmType', '')):
                    if disabled_val == 'TRUE':
                        indiv_enable = "Enabled"
                    elif disabled_val == 'FALSE':
                        indiv_enable = "Disabled"
                    else:
                        indiv_enable = "{n/a}"
                else:
                    # Analog alarms
                    if disabled_val == 'TRUE':
                        indiv_enable = "Enabled"
                    elif disabled_val == 'FALSE':
                        indiv_enable = "Disabled"
                    else:
                        indiv_enable = "{n/a}"
                
                # Clean limit value (remove commas, handle discrete)
                limit_value = ""
                if not self.is_discrete(param.get('alarmType', '')):
                    raw_limit = param.get('value', '')
                    if raw_limit and raw_limit not in ['~', '--------']:
                        limit_value = raw_limit.replace(',', '')
                
                # Build row based on client's PHA-Pro format
                phapro_format = self.config.get("phapro_headers", "FLNG")
                alarm_type = param.get('alarmType', '')
                
                if phapro_format == "HFS":
                    # HF Sinclair 43-column format (matches Tag_Import template)
                    row = [
                        tag['unit'] if is_first_tag_for_unit and is_first_alarm_for_tag else "",  # 1. Unit
                        tag['tag_name'] if is_first_alarm_for_tag else "",  # 2. Starting Tag Name
                        tag['tag_name'] if is_first_alarm_for_tag else "",  # 3. New Tag Name
                        tag['desc'] if is_first_alarm_for_tag else "",  # 4. Old Tag Description
                        tag['desc'] if is_first_alarm_for_tag else "",  # 5. New Tag Description
                        tag['pid'] if is_first_alarm_for_tag else "",  # 6. P&ID
                        tag['range_min'] if is_first_alarm_for_tag else "",  # 7. Range Min
                        tag['range_max'] if is_first_alarm_for_tag else "",  # 8. Range Max
                        tag['eng_units'] if is_first_alarm_for_tag else "",  # 9. Engineering Units
                        tag_source if is_first_alarm_for_tag else "",  # 10. Tag Source
                        f"Point Type = {tag['point_type']}" if is_first_alarm_for_tag and tag['point_type'] else "" if not is_first_alarm_for_tag else "",  # 11. Rationalization (Tag) Comment
                        "Enabled" if is_first_alarm_for_tag else "",  # 12. Old Tag Enable Status
                        "Enabled" if is_first_alarm_for_tag else "",  # 13. New Tag Enable Status
                        alarm_type,  # 14. Starting Alarm Type
                        alarm_type,  # 15. New Alarm Type
                        indiv_enable,  # 16. Old Alarm Enable Status
                        indiv_enable,  # 17. New Alarm Enable Status
                        priority_code,  # 18. Old (BPCS) Priority
                        priority_code,  # 19. New (BPCS) Priority
                        limit_value,  # 20. Old Limit
                        limit_value,  # 21. New Limit
                        self._clean_value(param.get('DeadBandValue', '')),  # 22. Old Deadband
                        self._clean_value(param.get('DeadBandValue', '')),  # 23. New Deadband
                        self._clean_value(param.get('DeadBandUnitValue', '')),  # 24. Old Deadband Units
                        self._clean_value(param.get('DeadBandUnitValue', '')),  # 25. New Deadband Units
                        self._clean_value(param.get('OnDelayValue', '')),  # 26. Old On-Delay Time
                        self._clean_value(param.get('OnDelayValue', '')),  # 27. New On-Delay Time
                        self._clean_value(param.get('OffDelayValue', '')),  # 28. Old Off-Delay Time
                        self._clean_value(param.get('OffDelayValue', '')),  # 29. New Off-Delay Time
                        "Not Started_x",  # 30. Rationalization Status
                        alarm_status,  # 31. Alarm Status
                        "",  # 32. Rationalization (Alarm) Comment
                        "",  # 33. Alarm Class
                        param.get('PurposeOfAlarm', '~') or "~",  # 34. Cause(s)
                        param.get('ConsequenceOfNoAction', '~') or "~",  # 35. Consequence(s)
                        param.get('BoardOperator', '~') or "~",  # 36. Inside Action(s)
                        param.get('FieldOperator', '~') or "~",  # 37. Outside Action(s)
                        "",  # 38. Escalation
                        "",  # 39. Limit Owner
                        "",  # 40. Personnel
                        "",  # 41. Public or Environment
                        self.map_severity(param.get('consequence', '')),  # 42. Costs / Production
                        param.get('TimeToRespond', '') or "",  # 43. Maximum Time to Resolve
                    ]
                else:
                    # FLNG 45-column format (default)
                    row = [
                        tag['unit'] if is_first_tag_for_unit and is_first_alarm_for_tag else "",
                        tag['tag_name'] if is_first_alarm_for_tag else "",
                        tag['desc'] or "~" if is_first_alarm_for_tag else "",
                        tag['desc'] or "~" if is_first_alarm_for_tag else "",
                        tag['pid'] if is_first_alarm_for_tag else "",
                        tag['range_min'] if is_first_alarm_for_tag else "",
                        tag['range_max'] if is_first_alarm_for_tag else "",
                        tag['eng_units'] or "~" if is_first_alarm_for_tag else "",
                        tag_source if is_first_alarm_for_tag else "",
                        f"Point Type = {tag['point_type']}" if is_first_alarm_for_tag and tag['point_type'] else "" if not is_first_alarm_for_tag else "",
                        "Enabled" if is_first_alarm_for_tag else "",
                        "Enabled" if is_first_alarm_for_tag else "",
                        alarm_type,
                        indiv_enable,
                        indiv_enable,
                        priority_code,
                        priority_code,
                        limit_value,
                        limit_value,
                        self._clean_value(param.get('DeadBandValue', '')),
                        self._clean_value(param.get('DeadBandValue', '')),
                        self._clean_value(param.get('DeadBandUnitValue', '')),
                        self._clean_value(param.get('DeadBandUnitValue', '')),
                        self._clean_value(param.get('OnDelayValue', '')),
                        self._clean_value(param.get('OnDelayValue', '')),
                        self._clean_value(param.get('OffDelayValue', '')),
                        self._clean_value(param.get('OffDelayValue', '')),
                        "Not Started_x",
                        alarm_status,
                        "", "", "", "", "",
                        param.get('PurposeOfAlarm', '~') or "~",
                        param.get('ConsequenceOfNoAction', '~') or "~",
                        param.get('BoardOperator', '~') or "~",
                        param.get('FieldOperator', '~') or "~",
                        "", "",
                        self.map_severity(param.get('consequence', '')),
                        "", "",
                        self.map_severity(param.get('consequence', '')),
                        param.get('TimeToRespond', '') or "",
                    ]
                
                rows.append(row)
                
                if is_first_alarm_for_tag:
                    last_unit = tag['unit']
                is_first_alarm_for_tag = False
        
        # Convert to CSV with Latin-1 encoding for DynAMo compatibility
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(self.get_phapro_headers())
        writer.writerows(rows)
        
        # Encode as Latin-1 bytes for proper download
        csv_string = output.getvalue()
        app_logger.info(f"Forward transform complete - tags: {self.stats['tags']}, alarms: {self.stats['alarms']}")
        return csv_string.encode('latin-1', errors='replace'), self.stats

    def transform_forward_abb(self, file_bytes: bytes) -> Tuple[str, Dict]:
        """Transform ABB Excel export to PHA-Pro format (23-column)."""
        tags = self.parse_abb_excel(file_bytes)
        
        rows = []
        self.stats = {"tags": 0, "alarms": 0, "units": set()}
        
        unit = self.config.get('unit_value', 'Line 1')
        self.stats["units"].add(unit)
        tag_source = self.config.get('default_source', 'ABB 800xA (DCS)')
        
        is_first_tag = True
        
        for tag in tags:
            self.stats["tags"] += 1
            is_first_alarm = True
            
            for alarm in tag['alarms']:
                self.stats["alarms"] += 1
                
                # Determine alarm status
                if alarm['enabled'] == 1:
                    alarm_status = "Alarm"
                else:
                    alarm_status = "None"
                
                # Format level (use empty for disabled or -9999999)
                level_val = alarm['level']
                if level_val == -9999999 or level_val == -9999999.0:
                    level_str = "-9999999"
                else:
                    level_str = str(int(level_val)) if level_val == int(level_val) else str(level_val)
                
                row = [
                    unit if is_first_tag and is_first_alarm else "",  # Unit
                    tag['tag_name'] if is_first_alarm else "",  # Starting Tag Name
                    tag['tag_name'] if is_first_alarm else "",  # New Tag Name
                    tag['description'] if is_first_alarm else "",  # Old Tag Description
                    tag['description'] if is_first_alarm else "",  # New Tag Description
                    tag_source if is_first_alarm else "",  # Tag Source
                    f"Tag Type = Analog Input" if is_first_alarm else "",  # Rationalization (Tag) Comment
                    "-9999999" if is_first_alarm else "",  # Range Min
                    "-9999999" if is_first_alarm else "",  # Range Max
                    "" if is_first_alarm else "",  # Engineering Units
                    alarm['type'],  # Starting Alarm Type
                    alarm['type'],  # New Alarm Type
                    str(alarm['enabled']),  # Old Alarm Enable Status
                    str(alarm['enabled']),  # New Alarm Enable Status
                    str(alarm['severity']),  # Old Alarm Severity
                    str(alarm['severity']),  # New Alarm Severity
                    level_str,  # Old Limit
                    level_str,  # New Limit
                    str(alarm['priority']),  # Old (BPCS) Priority
                    str(alarm['priority']),  # New (BPCS) Priority
                    "Not Started_x",  # Rationalization Status
                    alarm_status,  # Alarm Status
                    "",  # Rationalization (Alarm) Comment
                ]
                
                rows.append(row)
                
                if is_first_tag and is_first_alarm:
                    is_first_tag = False
                is_first_alarm = False
        
        # Convert to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(self.ABB_PHAPRO_HEADERS)
        writer.writerows(rows)
        
        return output.getvalue(), self.stats
    
    def transform_reverse_abb(self, file_content: str) -> Tuple[str, Dict]:
        """Transform PHA-Pro export back to ABB 8-column format."""
        lines = file_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        reader = csv.reader(lines)
        
        headers = next(reader)
        col_map = {h.strip(): i for i, h in enumerate(headers)}
        
        rows = []
        self.stats = {"tags": 0, "alarms": 0, "units": set()}
        seen_tags = set()
        
        last_tag_name = ""
        last_description = ""
        
        for row in reader:
            if not row or not any(row):
                continue
            
            # Get tag name (propagate from previous row if blank)
            tag_name_col = col_map.get('New Tag Name', col_map.get('Starting Tag Name', col_map.get('Tag Name', 1)))
            tag_name = row[tag_name_col].strip() if tag_name_col < len(row) else ""
            
            is_first_alarm_for_tag = False
            if tag_name:
                if tag_name != last_tag_name:
                    is_first_alarm_for_tag = True
                last_tag_name = tag_name
                # Also get description
                desc_col = col_map.get('New Tag Description', col_map.get('Old Tag Description', 4))
                if desc_col < len(row) and row[desc_col].strip():
                    last_description = row[desc_col].strip()
            else:
                tag_name = last_tag_name
            
            if tag_name not in seen_tags:
                seen_tags.add(tag_name)
                self.stats["tags"] += 1
            
            # Get alarm type
            alarm_type_col = col_map.get('New Alarm Type', col_map.get('Starting Alarm Type', 11))
            alarm_type = row[alarm_type_col].strip() if alarm_type_col < len(row) else ""
            
            if not alarm_type:
                continue
            
            self.stats["alarms"] += 1
            
            # Get other fields
            def get_col(name, default=""):
                idx = col_map.get(name)
                if idx is not None and idx < len(row):
                    val = row[idx].strip()
                    return val if val else default
                return default
            
            enable_status = get_col('New Alarm Enable Status', '0')
            new_limit = get_col('New Limit', '-9999999')
            new_priority = get_col('New (BPCS) Priority', '3')
            new_severity = get_col('New Alarm Severity', '1')
            alarm_comment = get_col('Rationalization (Alarm) Comment', '')
            
            # Build consolidated notes (similar to ABB format)
            notes = f"Cause:   Consequence:   Actions: {alarm_comment}" if alarm_comment else "Cause:   Consequence:   Actions: "
            
            output_row = [
                tag_name if is_first_alarm_for_tag else "",  # Tag Name (only on first alarm)
                last_description if is_first_alarm_for_tag else "",  # New Tag Description
                alarm_type,  # New Alarm Type
                enable_status,  # New Alarm Enable Status
                new_limit,  # New Limit
                new_priority,  # New Priority
                new_severity,  # New Alarm Severity Level
                notes,  # ABB Consolidated Notes
            ]
            
            rows.append(output_row)
        
        # Convert to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(self.ABB_RETURN_HEADERS)
        writer.writerows(rows)
        
        return output.getvalue(), self.stats

    def transform_reverse(self, file_content: str, source_data: Dict = None, selected_modes: List[str] = None) -> Tuple[str, Dict]:
        """
        Transform PHA-Pro export back to DynAMo format.
        
        This performs a MERGE operation:
        - Preserves all original DynAMo values for non-edited columns
        - Updates only specific columns from PHA-Pro rationalization results
        
        Columns UPDATED from PHA-Pro:
        - H (value/limit) <- New Limit
        - K (priorityValue) <- New (BPCS) Priority
        - M (consequence) <- Max Severity
        - N (TimeToRespond) <- Allowable Time to Respond / TTR Range
        - Q (Purpose of Alarm) <- Cause(s)
        - R (Consequence of No Action) <- Consequence(s)
        - S (Board Operator) <- Inside Action(s)
        - T (Field Operator) <- Outside Action(s)
        - Z (DisabledValue) <- New Individual Alarm Enable Status or derived from Rationalization Status
        
        All other columns preserved from original DynAMo file.
        """
        source_rows = len(source_data.get('rows', [])) if source_data else 0
        app_logger.info(f"Reverse transform started - client: {self.client_id}, source rows: {source_rows}")

        # Parse PHA-Pro file
        lines = file_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        reader = csv.reader(lines)
        
        headers = next(reader)
        col_map = {h.strip(): i for i, h in enumerate(headers)}
        
        # Validate required columns
        missing_columns = self.validate_phapro_columns(col_map)
        if missing_columns:
            raise ValueError(f"MISSING_COLUMNS:{','.join(missing_columns)}")
        
        # Determine which format we're working with
        phapro_format = self.config.get("phapro_headers", "FLNG")
        
        # Build lookup of PHA-Pro changes keyed by (tag_name, alarm_type)
        pha_changes = {}
        last_tag_name = ""
        last_tag_source = ""
        
        for row in reader:
            if not row or not any(row):
                continue
            
            # Get tag name - try multiple possible column names
            tag_name = self._get_col_flexible(col_map, row, ['Tag Name', 'New Tag Name', 'Starting Tag Name'], "")
            if tag_name:
                last_tag_name = tag_name
                # Get tag source if available
                tag_source = self._get_col_flexible(col_map, row, ['Tag Source'], "")
                if tag_source:
                    last_tag_source = tag_source
            else:
                tag_name = last_tag_name
            
            # Get alarm type - try multiple possible column names
            alarm_type = self._get_col_flexible(col_map, row, ['Alarm Type', 'New Alarm Type', 'Starting Alarm Type'], "")
            
            if not alarm_type:
                continue
            
            # Helper to get column value with fallbacks
            def get_col_flex(names, default=""):
                return self._get_col_flexible(col_map, row, names if isinstance(names, list) else [names], default)
            
            # Get Rationalization Status to derive alarm enable status for HFS
            rat_status = get_col_flex(['Rationalization Status'], '')
            
            # Derive enable status from Rationalization Status if no explicit enable column
            # Note: DisabledValue semantics:
            #   - DisabledValue = True  -> Alarm IS disabled (disabled flag is set)
            #   - DisabledValue = False -> Alarm is NOT disabled (enabled)
            enable_status = get_col_flex(['New Individual Alarm Enable Status', 'New Alarm Enable Status'], '')
            if not enable_status and rat_status:
                # HFS: "Deleted" means alarm should be disabled -> DisabledValue = True
                #      Other statuses mean alarm should be enabled -> DisabledValue = False
                if rat_status.lower() == 'deleted':
                    enable_status = 'True'  # Disabled flag is TRUE (alarm is disabled)
                else:
                    enable_status = 'False'  # Disabled flag is FALSE (alarm is enabled)
            
            # Store PHA-Pro values for this tag/alarm combination
            # Using flexible column name lookups
            pha_changes[(tag_name, alarm_type)] = {
                'new_limit': get_col_flex(['New Limit'], ''),
                'new_priority': get_col_flex(['New Priority', 'New (BPCS) Priority'], ''),
                'max_severity': get_col_flex(['Max Severity'], ''),
                'ttr': get_col_flex(['TTR Range', 'Max Time to Resolve', 'Allowable Time to Respond'], '~'),
                'causes': get_col_flex(['Cause(s)'], '~'),
                'consequences': get_col_flex(['Consequence(s)'], '~'),
                'inside_actions': get_col_flex(['Inside Action(s)'], '~'),
                'outside_actions': get_col_flex(['Outside Action(s)'], '~'),
                'new_enable_status': enable_status,
                'tag_source': last_tag_source,
                'rationalization_status': rat_status,
            }
        
        # If no source data provided, we can't do a proper merge
        if not source_data or 'rows' not in source_data:
            raise ValueError("Original DynAMo export file is required for reverse transformation. Please upload the original file.")
        
        # Process each row from original DynAMo file
        # IMPORTANT: Only output rows with valid mode (NORMAL or empty if empty_mode_is_valid)
        rows = []
        self.stats = {"tags": 0, "alarms": 0, "units": set(), "updated": 0, "not_found": 0, "skipped_modes": 0}
        seen_tags = set()
        seen_keys = set()  # Track (tag, alarm_type) to avoid duplicates
        
        # Check if empty mode is valid for this client
        empty_mode_is_valid = self.config.get("empty_mode_is_valid", False)
        
        for original_row in source_data['rows']:
            # Original row should have at least the key columns
            if len(original_row) < 6:
                continue
            
            # Check if this is a _Parameter row
            if original_row[0] != "_Variable" or original_row[2] != "_Parameter":
                continue
            
            tag_name = original_row[1]
            mode = original_row[3] if len(original_row) > 3 else ""
            alarm_type = original_row[5]
            
            # FILTER: Only include rows with valid mode
            # If selected_modes provided, use those; otherwise fall back to config behavior
            mode_upper = mode.upper().strip()
            if selected_modes is not None:
                # User explicitly selected modes
                selected_upper = [m.upper() for m in selected_modes if m != "(empty)"]
                allow_empty = "(empty)" in selected_modes
                if mode_upper in selected_upper:
                    pass  # Valid - mode is in user's selection
                elif allow_empty and mode.strip() == '':
                    pass  # Valid - empty mode and user selected (empty)
                else:
                    self.stats["skipped_modes"] += 1
                    continue
            else:
                # Legacy behavior
                if mode_upper == "NORMAL":
                    pass  # Valid
                elif mode_upper == "" and empty_mode_is_valid:
                    pass  # Valid for HFS
                else:
                    self.stats["skipped_modes"] += 1
                    continue
            
            # Skip placeholder alarm types (~ or empty)
            if not alarm_type or alarm_type.strip() in ['~', '-', '']:
                continue
            
            # Skip duplicate (tag, alarm_type) combinations
            key = (tag_name, alarm_type)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            
            if tag_name not in seen_tags:
                seen_tags.add(tag_name)
                self.stats["tags"] += 1
            
            self.stats["alarms"] += 1
            
            # Start with a copy of the original row (preserve all 42 columns)
            # Ensure row has all 42 columns
            output_row = list(original_row)
            while len(output_row) < 42:
                output_row.append("")
            
            # Clean commas from numeric fields (OnDelay, OffDelay, DeadBand values)
            # These are columns: AF (31), AI (34), AL (37)
            numeric_cols = [31, 34, 37]
            for col_idx in numeric_cols:
                if col_idx < len(output_row) and output_row[col_idx]:
                    val = output_row[col_idx]
                    # Remove commas from numbers like "1,500" -> "1500"
                    if ',' in val:
                        try:
                            # Check if it's a number with commas
                            cleaned = val.replace(',', '')
                            float(cleaned)  # Verify it's numeric
                            output_row[col_idx] = cleaned
                        except ValueError:
                            pass  # Not a number, keep as-is
            
            # Add apostrophe to _Variable for import
            output_row[0] = "'_Variable"
            
            # Look up PHA-Pro changes for this tag/alarm
            if key in pha_changes:
                changes = pha_changes[key]
                self.stats["updated"] += 1
                
                # Determine enforcement based on tag source
                tag_source = changes.get('tag_source', '')
                is_sm = "safety manager" in tag_source.lower()
                enforcement = "R" if is_sm else "M"
                
                # --- UPDATE COLUMN H (index 7): value/limit ---
                new_limit = changes['new_limit']
                at_lower = alarm_type.lower()
                
                # Handle special values
                if new_limit in ['{n/a}', '(n/a)', 'n/a', '{N/A}', '(N/A)', 'N/A']:
                    new_limit = '~'
                
                # Determine value based on alarm type
                # For discrete alarms: use ~
                # For analog alarms with no valid limit: use --------
                # For analog alarms with valid limit: use the limit value
                if self.is_discrete(alarm_type):
                    # Discrete alarms: keep ~ (not --------)
                    value = "~"
                elif "significant change" in at_lower:
                    # Significant change alarms: use --------
                    value = "--------"
                elif new_limit and new_limit not in ["~", "", "-9999999"]:
                    # Has a valid limit value
                    # Strip trailing zeros from decimal numbers (0.500 -> 0.5)
                    value = new_limit.replace(',', '')  # Remove commas first
                    try:
                        # Try to parse as float and format without trailing zeros
                        num = float(value)
                        if num == int(num):
                            value = str(int(num))  # Whole number
                        else:
                            # Format float, strip trailing zeros
                            value = f"{num:g}"
                    except ValueError:
                        # Not a number, keep as-is
                        value = new_limit
                else:
                    # No valid limit (empty, ~, or -9999999) - use --------
                    # This covers Advisory Deviation, Deviation Low, Accumulator deviation, etc.
                    value = "--------"
                    value = "--------"
                
                output_row[7] = value  # Column H: value
                
                # Update enforcement for value ONLY if alarm name exists
                if output_row[6]:  # alarmName exists
                    output_row[8] = enforcement  # Column I: enforcement
                
                # --- UPDATE COLUMN K (index 10): priorityValue ---
                new_priority = changes['new_priority']
                
                # Handle {n/a} priority
                if new_priority in ['{n/a}', '(n/a)', 'n/a', '{N/A}', '(N/A)', 'N/A']:
                    priority_value = '~'
                else:
                    # Map priority code to DynAMo value
                    priority_map = {
                        'U': 'Urgent', 'URGENT': 'Urgent',
                        'C': 'Critical', 'CRITICAL': 'Critical',
                        'H': 'High', 'HIGH': 'High',
                        'M': 'Medium', 'MEDIUM': 'Medium',
                        'L': 'Low', 'LOW': 'Low',
                        'J': 'Journal', 'JOURNAL': 'Journal',
                        'JO': 'Journal', 
                        'N': 'None', 'NONE': 'None',
                        'NA': 'NOACTION', 'NOACTION': 'NOACTION',  # HFS specific
                        'E': 'EMERGNCY', 'EMERGNCY': 'EMERGNCY',  # HFS specific
                    }
                    priority_value = priority_map.get(new_priority.upper(), new_priority)
                
                output_row[10] = priority_value  # Column K: priorityValue
                
                # Column L (priorityEnforcement): Only update if original had a value
                # Don't add enforcement where it didn't exist (e.g., significant change, accumulator)
                original_l = original_row[11].strip() if len(original_row) > 11 else ""
                if original_l:
                    output_row[11] = enforcement  # Column L: priorityEnforcement
                # else keep original (empty)
                
                # --- UPDATE COLUMN M (index 12): consequence ---
                max_severity = changes['max_severity']
                if max_severity in ['A', 'B', 'C', 'D', 'E']:
                    output_row[12] = max_severity
                elif max_severity and max_severity.upper() in ['NONE', '(NONE)', '(N)', 'N']:
                    output_row[12] = '(None)'  # Standardize to (None)
                elif max_severity:
                    output_row[12] = max_severity
                # else keep original
                
                # --- UPDATE COLUMN N (index 13): TimeToRespond ---
                ttr = changes['ttr']
                if ttr and ttr != '~':
                    output_row[13] = ttr
                
                # --- UPDATE COLUMN Q (index 16): Purpose of Alarm (Cause) ---
                # Always update from PHA-Pro (even if ~)
                causes = changes['causes']
                if causes:
                    causes = self._fix_encoding(causes)
                    output_row[16] = causes
                
                # --- UPDATE COLUMN R (index 17): Consequence of No Action ---
                # Always update from PHA-Pro (even if ~)
                consequences = changes['consequences']
                if consequences:
                    consequences = self._fix_encoding(consequences)
                    output_row[17] = consequences
                
                # --- UPDATE COLUMN S (index 18): Board Operator (Inside Action) ---
                # Always update from PHA-Pro (even if ~)
                inside_actions = changes['inside_actions']
                if inside_actions:
                    inside_actions = self._fix_encoding(inside_actions)
                    output_row[18] = inside_actions
                
                # --- UPDATE COLUMN T (index 19): Field Operator (Outside Action) ---
                # Always update from PHA-Pro (even if ~)
                outside_actions = changes['outside_actions']
                if outside_actions:
                    outside_actions = self._fix_encoding(outside_actions)
                    output_row[19] = outside_actions
                
                # --- UPDATE COLUMN Z (index 25): DisabledValue ---
                new_enable_status = changes['new_enable_status']
                if new_enable_status:
                    # DisabledValue semantics in HFS DynAMo:
                    #   - DisabledValue = True  -> Alarm IS disabled (flag is set)
                    #   - DisabledValue = False -> Alarm is NOT disabled (enabled)
                    # Match original file capitalization (True/False not TRUE/FALSE)
                    enable_upper = new_enable_status.upper()
                    if enable_upper in ['TRUE', 'ENABLED', '1']:
                        output_row[25] = 'True'  # Alarm is disabled
                    elif enable_upper in ['FALSE', 'DISABLED', '0']:
                        output_row[25] = 'False'  # Alarm is enabled
                    # else keep original
            else:
                self.stats["not_found"] += 1
                # Keep original row as-is (just add apostrophe)
            
            rows.append(output_row)
        
        # Convert to CSV - DynAMo import format has NO header row
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        
        # Get the string result and encode as latin-1 bytes
        # This ensures non-breaking space (U+00A0) is encoded as single byte 0xa0
        # instead of UTF-8's two bytes 0xc2 0xa0, matching the manual file format
        result_str = output.getvalue()
        try:
            result = result_str.encode('latin-1')
        except UnicodeEncodeError:
            # If there are characters that can't be encoded as latin-1, use UTF-8
            result = result_str.encode('utf-8')

        app_logger.info(f"Reverse transform complete - tags: {self.stats.get('tags', 0)}, alarms: {self.stats.get('alarms', 0)}")
        return result, self.stats

    def generate_change_report(self, pha_content: str, source_data: Dict, selected_modes: List[str] = None) -> bytes:
        """
        Generate an Excel change report comparing original DynAMo values with PHA-Pro changes.
        
        Args:
            pha_content: The PHA-Pro export file content
            source_data: Dictionary with 'rows' containing original DynAMo _Parameter rows
            
        Returns:
            Excel file as bytes
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Parse PHA-Pro file to get changes
        lines = pha_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        reader = csv.reader(lines)
        rows_list = list(reader)
        
        if not rows_list:
            raise ValueError("PHA-Pro file is empty")
        
        header = rows_list[0]
        col_map = {col.strip(): i for i, col in enumerate(header)}
        
        # Build PHA changes lookup
        pha_changes = {}
        last_tag_name = ""
        last_tag_source = ""
        
        for row in rows_list[1:]:
            if not row or not any(row):
                continue
            
            tag_name_idx = col_map.get('Tag Name', 1)
            tag_name = row[tag_name_idx].strip() if tag_name_idx < len(row) else ""
            if tag_name:
                last_tag_name = tag_name
                tag_source_idx = col_map.get('Tag Source', 5)
                if tag_source_idx < len(row) and row[tag_source_idx].strip():
                    last_tag_source = row[tag_source_idx].strip()
            else:
                tag_name = last_tag_name
            
            alarm_type_idx = col_map.get('Alarm Type', 7)
            alarm_type = row[alarm_type_idx].strip() if alarm_type_idx < len(row) else ""
            
            if not alarm_type:
                continue
            
            def get_col(name, default=""):
                idx = col_map.get(name)
                if idx is not None and idx < len(row):
                    return row[idx].strip() or default
                return default
            
            pha_changes[(tag_name, alarm_type)] = {
                'new_limit': get_col('New Limit', ''),
                'new_priority': get_col('New Priority', ''),
                'max_severity': get_col('Max Severity', ''),
                'ttr': get_col('TTR Range', ''),
                'causes': get_col('Cause(s)', ''),
                'consequences': get_col('Consequence(s)', ''),
                'inside_actions': get_col('Inside Action(s)', ''),
                'outside_actions': get_col('Outside Action(s)', ''),
                'new_enable_status': get_col('New Individual Alarm Enable Status', ''),
                'tag_source': last_tag_source,
                'unit': get_col('Unit', ''),
            }
        
        # Build change records by comparing original with PHA changes
        change_records = []
        seen_keys = set()
        
        # Column indices in DynAMo file
        # H=7 (value), K=10 (priorityValue), M=12 (consequence), N=13 (TTR)
        # Q=16 (Purpose), R=17 (Consequence), S=18 (Board Op), T=19 (Field Op)
        # Z=25 (DisabledValue)
        
        for original_row in source_data.get('rows', []):
            if len(original_row) < 6:
                continue
            
            if original_row[0] not in ["_Variable", "'_Variable"] or original_row[2] != "_Parameter":
                continue
            
            tag_name = original_row[1]
            mode = original_row[3] if len(original_row) > 3 else ""
            alarm_type = original_row[5]
            
            # Filter by selected modes or default to NORMAL
            mode_upper = mode.upper().strip()
            if selected_modes is not None:
                selected_upper = [m.upper() for m in selected_modes if m != "(empty)"]
                allow_empty = "(empty)" in selected_modes
                if mode_upper not in selected_upper and not (allow_empty and mode.strip() == ''):
                    continue
            else:
                if mode_upper != "NORMAL":
                    continue
            
            key = (tag_name, alarm_type)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            
            if key not in pha_changes:
                continue
            
            changes = pha_changes[key]
            
            # Get original values
            orig_value = original_row[7].strip() if len(original_row) > 7 else ""
            orig_priority = original_row[10].strip() if len(original_row) > 10 else ""
            orig_consequence = original_row[12].strip() if len(original_row) > 12 else ""
            orig_ttr = original_row[13].strip() if len(original_row) > 13 else ""
            orig_purpose = original_row[16].strip() if len(original_row) > 16 else ""
            orig_conseq_action = original_row[17].strip() if len(original_row) > 17 else ""
            orig_board_op = original_row[18].strip() if len(original_row) > 18 else ""
            orig_field_op = original_row[19].strip() if len(original_row) > 19 else ""
            orig_disabled = original_row[25].strip() if len(original_row) > 25 else ""
            
            # Calculate new values (same logic as transform_reverse)
            new_limit = changes['new_limit']
            at_lower = alarm_type.lower()
            
            if new_limit in ['{n/a}', '(n/a)', 'n/a', '{N/A}', '(N/A)', 'N/A']:
                new_limit = '~'
            
            if self.is_discrete(alarm_type):
                new_value = "~"
            elif "significant change" in at_lower:
                new_value = "--------"
            elif new_limit and new_limit not in ["~", "", "-9999999"]:
                new_value = new_limit.replace(',', '')
                try:
                    num = float(new_value)
                    if num == int(num):
                        new_value = str(int(num))
                    else:
                        new_value = f"{num:g}"
                except ValueError:
                    new_value = new_limit
            else:
                new_value = "--------"
            
            # Priority mapping
            new_priority_raw = changes['new_priority']
            if new_priority_raw in ['{n/a}', '(n/a)', 'n/a', '{N/A}', '(N/A)', 'N/A']:
                new_priority = '~'
            else:
                priority_map = {
                    'U': 'Urgent', 'URGENT': 'Urgent', 'C': 'Critical', 'CRITICAL': 'Critical',
                    'H': 'High', 'HIGH': 'High', 'M': 'Medium', 'MEDIUM': 'Medium',
                    'L': 'Low', 'LOW': 'Low', 'J': 'Journal', 'JOURNAL': 'Journal',
                    'JO': 'Journal', 'N': 'None', 'NONE': 'None',
                }
                new_priority = priority_map.get(new_priority_raw.upper(), new_priority_raw) if new_priority_raw else ''
            
            # Consequence/severity
            max_severity = changes['max_severity']
            if max_severity in ['A', 'B', 'C', 'D', 'E']:
                new_consequence = max_severity
            elif max_severity and max_severity.upper() in ['NONE', '(NONE)', '(N)', 'N']:
                new_consequence = '(None)'
            else:
                new_consequence = max_severity or orig_consequence
            
            new_ttr = changes['ttr'] if changes['ttr'] and changes['ttr'] != '~' else orig_ttr
            new_purpose = changes['causes'] if changes['causes'] else orig_purpose
            new_conseq_action = changes['consequences'] if changes['consequences'] else orig_conseq_action
            new_board_op = changes['inside_actions'] if changes['inside_actions'] else orig_board_op
            new_field_op = changes['outside_actions'] if changes['outside_actions'] else orig_field_op
            
            # Disabled value
            new_enable_status = changes['new_enable_status']
            if new_enable_status:
                enable_upper = new_enable_status.upper()
                if enable_upper in ['TRUE', 'ENABLED', '1']:
                    new_disabled = 'TRUE'
                elif enable_upper in ['FALSE', 'DISABLED', '0']:
                    new_disabled = 'FALSE'
                else:
                    new_disabled = orig_disabled
            else:
                new_disabled = orig_disabled
            
            # Check if any field changed
            value_changed = orig_value != new_value
            priority_changed = orig_priority != new_priority
            consequence_changed = orig_consequence != new_consequence
            ttr_changed = orig_ttr != new_ttr
            purpose_changed = orig_purpose != new_purpose
            conseq_action_changed = orig_conseq_action != new_conseq_action
            board_op_changed = orig_board_op != new_board_op
            field_op_changed = orig_field_op != new_field_op
            disabled_changed = orig_disabled != new_disabled
            
            any_change = (value_changed or priority_changed or consequence_changed or 
                         ttr_changed or purpose_changed or conseq_action_changed or 
                         board_op_changed or field_op_changed or disabled_changed)
            
            if any_change:
                change_records.append({
                    'Unit': changes.get('unit', ''),
                    'Tag Name': tag_name,
                    'Alarm Type': alarm_type,
                    'Tag Source': changes.get('tag_source', ''),
                    # Value/Limit
                    'Original Limit': orig_value,
                    'New Limit': new_value,
                    'Limit Changed': 'âœ“' if value_changed else '',
                    # Priority
                    'Original Priority': orig_priority,
                    'New Priority': new_priority,
                    'Priority Changed': 'âœ“' if priority_changed else '',
                    # Consequence/Severity
                    'Original Severity': orig_consequence,
                    'New Severity': new_consequence,
                    'Severity Changed': 'âœ“' if consequence_changed else '',
                    # TTR
                    'Original TTR': orig_ttr,
                    'New TTR': new_ttr,
                    'TTR Changed': 'âœ“' if ttr_changed else '',
                    # Purpose/Cause
                    'Original Purpose': orig_purpose[:100] + '...' if len(orig_purpose) > 100 else orig_purpose,
                    'New Purpose': new_purpose[:100] + '...' if len(new_purpose) > 100 else new_purpose,
                    'Purpose Changed': 'âœ“' if purpose_changed else '',
                    # Consequence of No Action
                    'Original Consequence': orig_conseq_action[:100] + '...' if len(orig_conseq_action) > 100 else orig_conseq_action,
                    'New Consequence': new_conseq_action[:100] + '...' if len(new_conseq_action) > 100 else new_conseq_action,
                    'Consequence Changed': 'âœ“' if conseq_action_changed else '',
                    # Board Operator
                    'Original Board Op': orig_board_op[:100] + '...' if len(orig_board_op) > 100 else orig_board_op,
                    'New Board Op': new_board_op[:100] + '...' if len(new_board_op) > 100 else new_board_op,
                    'Board Op Changed': 'âœ“' if board_op_changed else '',
                    # Field Operator
                    'Original Field Op': orig_field_op[:100] + '...' if len(orig_field_op) > 100 else orig_field_op,
                    'New Field Op': new_field_op[:100] + '...' if len(new_field_op) > 100 else new_field_op,
                    'Field Op Changed': 'âœ“' if field_op_changed else '',
                    # Enabled/Disabled
                    'Original Enabled': orig_disabled,
                    'New Enabled': new_disabled,
                    'Enabled Changed': 'âœ“' if disabled_changed else '',
                })
        
        # Create DataFrame
        df = pd.DataFrame(change_records)
        
        # Create Excel workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Change Report"
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        change_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
        checkmark_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write header
        if len(df) > 0:
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Write data rows
            for row_idx, record in enumerate(df.to_dict('records'), 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Highlight "Changed" columns with checkmarks
                    if 'Changed' in header and value == 'âœ“':
                        cell.fill = checkmark_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Highlight "New" columns that have changes
                    if header.startswith('New ') and not header.endswith('Changed'):
                        # Check if corresponding Changed column has checkmark
                        base_name = header.replace('New ', '')
                        changed_col = f"{base_name} Changed"
                        if changed_col in record and record[changed_col] == 'âœ“':
                            cell.fill = change_fill
            
            # Adjust column widths
            column_widths = {
                'Unit': 12, 'Tag Name': 20, 'Alarm Type': 20, 'Tag Source': 25,
                'Original Limit': 12, 'New Limit': 12, 'Limit Changed': 8,
                'Original Priority': 12, 'New Priority': 12, 'Priority Changed': 8,
                'Original Severity': 12, 'New Severity': 12, 'Severity Changed': 8,
                'Original TTR': 18, 'New TTR': 18, 'TTR Changed': 8,
                'Original Purpose': 35, 'New Purpose': 35, 'Purpose Changed': 8,
                'Original Consequence': 35, 'New Consequence': 35, 'Consequence Changed': 8,
                'Original Board Op': 35, 'New Board Op': 35, 'Board Op Changed': 8,
                'Original Field Op': 35, 'New Field Op': 35, 'Field Op Changed': 8,
                'Original Enabled': 12, 'New Enabled': 12, 'Enabled Changed': 8,
            }
            
            for col_idx, header in enumerate(headers, 1):
                width = column_widths.get(header, 15)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        else:
            ws.cell(row=1, column=1, value="No changes detected")
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Change Report Summary").font = Font(bold=True, size=14)
        ws_summary.cell(row=3, column=1, value="Total Alarms with Changes:")
        ws_summary.cell(row=3, column=2, value=len(df))
        
        # Count changes by type
        if len(df) > 0:
            ws_summary.cell(row=5, column=1, value="Changes by Field:").font = Font(bold=True)
            change_cols = [col for col in df.columns if col.endswith('Changed')]
            row = 6
            for col in change_cols:
                count = (df[col] == 'âœ“').sum()
                if count > 0:
                    field_name = col.replace(' Changed', '')
                    ws_summary.cell(row=row, column=1, value=field_name)
                    ws_summary.cell(row=row, column=2, value=count)
                    row += 1
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    def generate_change_report_abb(self, pha_content: str, source_bytes: bytes) -> bytes:
        """
        Generate an Excel change report comparing original ABB values with PHA-Pro changes.

        Args:
            pha_content: The PHA-Pro export file content
            source_bytes: Original ABB Excel file as bytes

        Returns:
            Excel file as bytes
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Parse original ABB Excel to get original values
        original_tags = self.parse_abb_excel(source_bytes)
        original_lookup = {}
        for tag in original_tags:
            for alarm in tag['alarms']:
                key = (tag['tag_name'], alarm['type'])
                original_lookup[key] = {
                    'description': tag['description'],
                    'limit': alarm['level'],
                    'priority': alarm['priority'],
                    'severity': alarm['severity'],
                    'enabled': alarm['enabled'],
                }

        # Parse PHA-Pro file to get rationalized values
        lines = pha_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        reader = csv.reader(lines)
        rows_list = list(reader)

        if not rows_list:
            raise ValueError("PHA-Pro file is empty")

        header = rows_list[0]
        col_map = {col.strip(): i for i, col in enumerate(header)}

        # Build change records
        change_records = []
        last_tag_name = ""

        for row in rows_list[1:]:
            if not row or not any(row):
                continue

            # Get tag name (propagate from previous row if blank)
            tag_name_idx = col_map.get('New Tag Name', col_map.get('Starting Tag Name', col_map.get('Tag Name', 1)))
            tag_name = row[tag_name_idx].strip() if tag_name_idx < len(row) else ""
            if tag_name:
                last_tag_name = tag_name
            else:
                tag_name = last_tag_name

            # Get alarm type
            alarm_type_idx = col_map.get('New Alarm Type', col_map.get('Starting Alarm Type', 11))
            alarm_type = row[alarm_type_idx].strip() if alarm_type_idx < len(row) else ""

            if not alarm_type:
                continue

            key = (tag_name, alarm_type)
            original = original_lookup.get(key, {})

            def get_col(name, default=""):
                idx = col_map.get(name)
                if idx is not None and idx < len(row):
                    return row[idx].strip() or default
                return default

            # Get original values
            orig_limit = original.get('limit', '')
            orig_priority = original.get('priority', '')
            orig_severity = original.get('severity', '')
            orig_enabled = original.get('enabled', '')

            # Format original limit
            if orig_limit == -9999999 or orig_limit == -9999999.0:
                orig_limit_str = "-9999999"
            elif orig_limit:
                orig_limit_str = str(int(orig_limit)) if orig_limit == int(orig_limit) else str(orig_limit)
            else:
                orig_limit_str = ""

            # Get new values from PHA-Pro
            new_limit = get_col('New Limit', orig_limit_str)
            new_priority = get_col('New (BPCS) Priority', get_col('New Priority', str(orig_priority)))
            new_severity = get_col('New Alarm Severity', str(orig_severity))
            new_enabled = get_col('New Alarm Enable Status', str(orig_enabled))

            # Check for changes
            limit_changed = str(orig_limit_str) != str(new_limit)
            priority_changed = str(orig_priority) != str(new_priority)
            severity_changed = str(orig_severity) != str(new_severity)
            enabled_changed = str(orig_enabled) != str(new_enabled)

            any_change = limit_changed or priority_changed or severity_changed or enabled_changed

            if any_change:
                change_records.append({
                    'Tag Name': tag_name,
                    'Alarm Type': alarm_type,
                    'Original Limit': orig_limit_str,
                    'New Limit': new_limit,
                    'Limit Changed': '✓' if limit_changed else '',
                    'Original Priority': str(orig_priority),
                    'New Priority': new_priority,
                    'Priority Changed': '✓' if priority_changed else '',
                    'Original Severity': str(orig_severity),
                    'New Severity': new_severity,
                    'Severity Changed': '✓' if severity_changed else '',
                    'Original Enabled': str(orig_enabled),
                    'New Enabled': new_enabled,
                    'Enabled Changed': '✓' if enabled_changed else '',
                })

        # Create DataFrame
        df = pd.DataFrame(change_records)

        # Create Excel workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Change Report"

        # Define styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        change_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        checkmark_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header
        if len(df) > 0:
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Write data rows
            for row_idx, record in enumerate(df.to_dict('records'), 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

                    # Highlight "Changed" columns with checkmarks
                    if 'Changed' in header and value == '✓':
                        cell.fill = checkmark_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Highlight "New" columns that have changes
                    if header.startswith('New ') and not header.endswith('Changed'):
                        base_name = header.replace('New ', '')
                        changed_col = f"{base_name} Changed"
                        if changed_col in record and record[changed_col] == '✓':
                            cell.fill = change_fill

            # Adjust column widths
            column_widths = {
                'Tag Name': 25, 'Alarm Type': 18,
                'Original Limit': 12, 'New Limit': 12, 'Limit Changed': 8,
                'Original Priority': 12, 'New Priority': 12, 'Priority Changed': 8,
                'Original Severity': 12, 'New Severity': 12, 'Severity Changed': 8,
                'Original Enabled': 12, 'New Enabled': 12, 'Enabled Changed': 8,
            }
            for col_idx, header in enumerate(headers, 1):
                width = column_widths.get(header, 15)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

            # Freeze header row
            ws.freeze_panes = 'A2'
        else:
            ws.cell(row=1, column=1, value="No changes detected")

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="ABB Change Report Summary").font = Font(bold=True, size=14)
        ws_summary.cell(row=3, column=1, value="Total Alarms with Changes:")
        ws_summary.cell(row=3, column=2, value=len(df))

        # Count changes by type
        if len(df) > 0:
            ws_summary.cell(row=5, column=1, value="Changes by Field:").font = Font(bold=True)
            change_cols = [col for col in df.columns if col.endswith('Changed')]
            row = 6
            for col in change_cols:
                count = (df[col] == '✓').sum()
                if count > 0:
                    field_name = col.replace(' Changed', '')
                    ws_summary.cell(row=row, column=1, value=field_name)
                    ws_summary.cell(row=row, column=2, value=count)
                    row += 1

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def scan_for_units(file_content: str, client_id: str) -> Tuple[set, set, set]:
    """
    Scan a DynAMo file and extract available units using multiple methods.
    
    Returns:
        Tuple of (units_by_tag_prefix, units_by_asset_parent, units_by_asset_child)
        - units_by_tag_prefix: First digits of tag name (e.g., "17" from "17TI5879")
        - units_by_asset_parent: First level after U## (e.g., "17_FLARE" from /U17/17_FLARE/17H-2)
        - units_by_asset_child: Last level before tag (e.g., "17H-2" from /U17/17_FLARE/17H-2)
    """
    import re
    
    lines = file_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    reader = csv.reader(lines)
    
    units_by_prefix = set()
    units_by_asset_parent = set()
    units_by_asset_child = set()
    
    # Get config for unit extraction
    config = AlarmTransformer.get_client_configs().get(client_id, AlarmTransformer.get_client_configs()["flng"])
    unit_digits = config.get("unit_digits", 2)
    
    for row in reader:
        if not row or len(row) < 4:
            continue
        
        if row[0].strip() == "_Variable" and len(row) > 2:
            tag_name = row[1].strip()
            schema_type = row[2].strip()
            
            if schema_type == "_DCSVariable":
                # Extract unit from tag prefix
                unit = ""
                for ch in tag_name:
                    if ch.isdigit():
                        unit += ch
                        if len(unit) >= unit_digits:
                            break
                    elif unit:
                        break
                if unit:
                    units_by_prefix.add(unit)
                
                # Extract units from asset path
                asset_path = row[3] if len(row) > 3 else ""
                if asset_path:
                    # Parse asset path: /Assets/LQF/U17/17_FLARE/17H-2
                    # We want: parent = 17_FLARE, child = 17H-2
                    
                    # Find the U## level first
                    match = re.search(r'/U(\d+)/', asset_path, re.IGNORECASE)
                    if match:
                        # Get everything after /U##/
                        u_pos = match.end()
                        remaining = asset_path[u_pos:]
                        
                        # Split by /
                        parts = [p for p in remaining.split('/') if p]
                        
                        if len(parts) >= 1:
                            # Parent unit is first level after U##
                            units_by_asset_parent.add(parts[0])
                        
                        if len(parts) >= 2:
                            # Child unit is last level (if different from parent)
                            child = parts[-1]
                            if child != parts[0]:
                                units_by_asset_child.add(child)
                        elif len(parts) == 1:
                            # No child, just parent (tag is directly under parent)
                            pass
    
    return units_by_prefix, units_by_asset_parent, units_by_asset_child


def scan_for_modes(file_content: str) -> set:
    """
    Scan a DynAMo file and extract all unique mode values from _Parameter rows.

    Args:
        file_content: The decoded CSV file content

    Returns:
        Set of unique mode strings found (e.g., {"NORMAL", "Base", "IMPORT"})
    """
    lines = file_content.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    reader = csv.reader(lines)

    modes = set()

    for row in reader:
        if not row or len(row) < 4:
            continue

        if row[0].strip() == "_Variable" and len(row) > 2:
            schema_type = row[2].strip()

            if schema_type == "_Parameter":
                mode = row[3].strip() if len(row) > 3 else ""
                if mode:
                    modes.add(mode)
                else:
                    modes.add("(empty)")

    return modes



# ============================================================
# STREAMLIT UI
# ============================================================

def main():
    # Check authentication first
    if not check_password():
        return
    
    # Sidebar first to get client selection
    with st.sidebar:
        # Logout button at top of sidebar - small and subtle
        st.markdown(
            f'<div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;">'
            f'<span style="font-size: 0.9rem;">👤 {st.session_state.get("username", "User")}</span>'
            f'</div>',
            unsafe_allow_html=True
        )
        if st.button("Logout", key="logout_btn", type="secondary"):
            st.session_state["authenticated"] = False
            st.session_state["username"] = None
            st.rerun()
        
        st.markdown("---")
        st.markdown("### ⚙️ Configuration")

        # Show config warnings if any
        if _CONFIG_WARNINGS:
            errors = [w for w in _CONFIG_WARNINGS if w['level'] == 'error']
            warnings = [w for w in _CONFIG_WARNINGS if w['level'] == 'warning']

            if errors:
                with st.expander(f"⛔ {len(errors)} Config Error(s)", expanded=True):
                    for err in errors:
                        st.error(f"**{err['client']}**: {err['message']}")

            if warnings:
                with st.expander(f"⚠️ {len(warnings)} Config Warning(s)"):
                    for warn in warnings:
                        st.warning(f"**{warn['client']}**: {warn['message']}")

        # Build client options from configs
        client_options = {
            client_id: config["name"] 
            for client_id, config in AlarmTransformer.get_client_configs().items()
        }
        
        selected_client = st.selectbox(
            "Select Client Profile",
            options=list(client_options.keys()),
            format_func=lambda x: client_options[x],
            help="Choose the client configuration for tag source rules and mappings"
        )

        # Get available areas for this client
        area_options = AlarmTransformer.get_client_areas(selected_client)

        # Area/Unit dropdown
        if area_options:
            selected_area = st.selectbox(
                "Select Unit/Area",
                options=list(area_options.keys()),
                format_func=lambda x: area_options[x],
                help="Choose the specific unit or area within this client site"
            )
        else:
            selected_area = None

        # Track client/area changes to clear file uploads
        # Initialize session state for file uploader reset
        if 'file_uploader_key' not in st.session_state:
            st.session_state.file_uploader_key = 0
        if 'previous_client' not in st.session_state:
            st.session_state.previous_client = selected_client
        if 'previous_area' not in st.session_state:
            st.session_state.previous_area = selected_area

        # Check if client or area changed - if so, reset file uploaders
        if (st.session_state.previous_client != selected_client or
            st.session_state.previous_area != selected_area):
            st.session_state.file_uploader_key += 1
            st.session_state.previous_client = selected_client
            st.session_state.previous_area = selected_area
        
        # Get the selected client's config for dynamic labels
        client_config = AlarmTransformer.get_client_configs().get(selected_client, {})
        dcs_name = client_config.get("dcs_name", "DCS")
        pha_tool = client_config.get("pha_tool", "PHA-Pro")
        
        # Request New Client/Area button
        import urllib.parse as sidebar_urllib
        current_area_name = area_options.get(selected_area, "N/A") if selected_area else "N/A"
        request_subject = sidebar_urllib.quote("New Client/Area Request - Alarm Platform")
        request_body = sidebar_urllib.quote(f"Hi Greg,\n\nI need a new client or unit/area added.\n\nCurrent: {client_options.get(selected_client, 'Unknown')}\nArea: {current_area_name}\n\nNew Request:\n- Client/Site: \n- Unit/Area: \n- DCS System: \n\nThanks")
        request_link = f"mailto:greg.pajak@aesolutions.com?subject={request_subject}&body={request_body}"
        st.markdown(
            f'<a href="{request_link}" style="text-decoration: none;">'
            f'<div style="background-color: #2d5a87; color: white; padding: 8px 12px; border-radius: 5px; text-align: center; font-size: 0.85rem; margin-top: 10px;">'
            f'📧 Request New Client/Area'
            f'</div></a>',
            unsafe_allow_html=True
        )
        st.caption("Don\'t see your client or unit? Click to request.")
        
        st.markdown("---")
        
        # Transformation direction with generic labels
        direction = st.radio(
            "Transformation Direction",
            options=["forward", "reverse"],
            format_func=lambda x: "Alarm Database → PHA-Pro" if x == "forward" else "PHA-Pro → Alarm Database",
            help="Forward: Create PHA-Pro import from alarm database export\nReverse: Create alarm database import from PHA-Pro export"
        )

        st.markdown("---")

        # Help section
        with st.expander("ℹ️ How to Use"):
            if client_config.get("parser", "dynamo") == "dynamo":
                st.markdown(f"""
                **Forward Transformation ({dcs_name} → {pha_tool})**
                1. Export your alarm database from {dcs_name} as CSV
                2. Upload the CSV file below
                3. Select unit extraction method (FLNG only):
                   - *Tag Prefix*: First digits of tag (e.g., "17" from "17TI5879")
                   - *Asset Parent*: Consolidated units (e.g., "17_FLARE")
                   - *Asset Child*: Detailed units (e.g., "17H-2")
                4. **Verify detected columns** - check the checkbox to confirm
                5. Click Transform
                6. Download the {pha_tool} import file
                7. Review output before importing to {pha_tool}
                
                **Reverse Transformation ({pha_tool} → {dcs_name})**
                1. Export from {pha_tool} MADB as CSV
                2. Upload the {pha_tool} export file
                3. **Upload the original {dcs_name} export** (required)
                4. Click Transform
                5. Download the {dcs_name} import file
                6. **Optional:** Download Change Report (Excel)
                
                **Key Mappings (Reverse):**
                - Rationalization Status "Deleted" → DisabledValue = True
                - Priority NA → NOACTION, E → EMERGNCY (HFS)
                - Priority J → Journal, H → High, etc.
                """)
            else:
                st.markdown(f"""
                **Forward Transformation ({dcs_name} → {pha_tool})**
                1. Export your alarm database from {dcs_name} as Excel
                2. Upload the Excel file below
                3. Click Transform
                4. Download the {pha_tool} import file
                
                **Reverse Transformation ({pha_tool} → {dcs_name})**
                1. Export from {pha_tool} MADB as CSV
                2. Upload the CSV file below
                3. Click Transform
                4. Download the {dcs_name} return file
                """)
        
        st.markdown("---")
        st.markdown("### 📊 About")
        st.markdown(f"""
        **Version:** 3.26  
        **Client:** {client_options.get(selected_client, 'Unknown')}  
        **Last Updated:** {datetime.now().strftime('%Y-%m-%d')}
        """)
        
        with st.expander("📝 Version History"):
            st.markdown("""
            **v3.22** - Jan 2026
            - Fixed Unit column: only shows on first row of each unit group (not every tag)
            
            **v3.21** - Jan 2026
            - Fixed encoding output (Latin-1 bytes for proper °F display)
            - Comma stripping in delay values (1,500 → 1500)
            
            **v3.19** - Jan 2026
            - Enhanced unit extraction: Tag Prefix, Asset Parent, Asset Child options
            - Asset Parent gives consolidated units (17_FLARE, 17_FGS, etc.)
            - Asset Child gives detailed units (17H-2, 17IB-02, etc.)
            
            **v3.18** - Jan 2026
            - Forward transform: Mode filtering (NORMAL only)
            - Forward transform: Full unit from _DCS[10]
            - Forward transform: Engineering units from _DCS[3]
            - Forward transform: Range Max comma removal
            - Forward transform: P&ID "UNKNOWN" default
            - Forward transform: {n/a} for discrete alarm enable status
            - P&ID review warning after forward transform
            
            **v3.17** - Jan 2026
            - Fixed UI timing (spinner completes before success message)
            - Change Report Excel export with formatting
            - 100% validation match with manual process
            
            **v3.15** - Jan 2026
            - Latin-1 encoding for DynAMo compatibility
            - Expanded discrete alarm type detection
            - Fixed value logic for deviation alarms
            
            **v3.8** - Jan 2026
            - Mode filtering (NORMAL only)
            - Decimal formatting (strip trailing zeros)
            - Skipped modes explanation UI
            
            **v3.1** - Jan 2026
            - Dynamic header and descriptions
            - Fixed Output Format column counts
            
            **v3.0** - Jan 2026
            - Added ABB 800xA support
            - Rio Tinto - Bessemer City client
            - Excel (.xlsx) input for ABB
            
            **v2.2** - Jan 2026  
            - Missing column validation
            - HF Sinclair - Artesia client
            - Dynamic radio button labels
            
            **v2.1** - Jan 2026
            - Unit detection on file upload
            - Tag Prefix / Asset Path / Both options
            
            **v2.0** - Jan 2026
            - Severity mapping fix (MINOR→D, etc.)
            - OnDelay, OffDelay, Deadband extraction
            - Freeport LNG client
            
            **v1.0** - Jan 2026
            - Initial release
            - DynAMo â†” PHA-Pro transformation
            """)
        
        # Bug/Feature Report Button
        st.markdown("---")
        st.markdown("### 🐛 Report Issue")
        
        report_type = st.selectbox(
            "What would you like to report?",
            ["Bug / Something broken", "Feature request", "Question / Other"],
            key="report_type"
        )
        
        # Build email subject and body based on report type
        subject_map = {
            "Bug / Something broken": "Bug Report - Alarm Rationalization Platform",
            "Feature request": "Feature Request - Alarm Rationalization Platform",
            "Question / Other": "Question - Alarm Rationalization Platform"
        }
        
        body_map = {
            "Bug / Something broken": f"""Hi Greg,

I found an issue with the Alarm Rationalization Platform.

CLIENT: {client_options.get(selected_client, 'Unknown')}
DIRECTION: {direction}
VERSION: 3.17

DESCRIPTION OF ISSUE:
[Describe what happened]

STEPS TO REPRODUCE:
1. 
2. 
3. 

EXPECTED BEHAVIOR:
[What should have happened]

ACTUAL BEHAVIOR:
[What actually happened]

Please attach any relevant files or screenshots.

Thanks,
{st.session_state.get('username', '[Your name]')}""",
            
            "Feature request": f"""Hi Greg,

I have a feature suggestion for the Alarm Rationalization Platform.

CLIENT: {client_options.get(selected_client, 'Unknown')}
VERSION: 3.3

FEATURE DESCRIPTION:
[Describe the feature you'd like]

WHY IT WOULD BE HELPFUL:
[Explain the use case]

Thanks,
{st.session_state.get('username', '[Your name]')}""",
            
            "Question / Other": f"""Hi Greg,

I have a question about the Alarm Rationalization Platform.

CLIENT: {client_options.get(selected_client, 'Unknown')}
VERSION: 3.3

QUESTION:
[Your question here]

Thanks,
{st.session_state.get('username', '[Your name]')}"""
        }
        
        import urllib.parse
        subject = urllib.parse.quote(subject_map[report_type])
        body = urllib.parse.quote(body_map[report_type])
        email_link = f"mailto:greg.pajak@aesolutions.com?subject={subject}&body={body}"
        
        st.markdown(
            f'<a href="{email_link}" target="_blank">'
            f'<button style="width:100%; padding:10px; background-color:#4a6fa5; color:white; border:none; border-radius:5px; cursor:pointer;">'
            f'📧 Open Email to Report'
            f'</button></a>',
            unsafe_allow_html=True
        )
        st.caption("Attach files/screenshots in your email client")

        # Session History Section
        st.markdown("---")
        st.markdown("### 📜 Session History")
        history = get_history()
        if history:
            st.caption(f"{len(history)} transformation(s)")
            with st.expander("View History", expanded=False):
                for entry in reversed(history):
                    st.markdown(f"✅ **{entry['input_file']}**  \n↳ {entry['direction']} | {entry['tags']} tags | {entry['alarms']} alarms  \n<small>{entry['timestamp']}</small>", unsafe_allow_html=True)
                    st.download_button(f"⬇️ {entry['output_filename']}", entry['output_data'], entry['output_filename'], "text/csv", key=f"hist_{entry['id']}", use_container_width=True)
                    st.markdown("---")
                if st.button("🗑️ Clear History", key="clear_hist"):
                    clear_history()
                    st.rerun()
        else:
            st.caption("No transformations yet")

        # Debug Logs Section
        with st.expander("🔍 Debug Logs", expanded=False):
            logs = st.session_state.get('app_logs', [])
            if logs:
                st.caption(f"{len(logs)} entries")
                for log in reversed(logs[-30:]):
                    color = {'INFO': 'green', 'WARNING': 'orange', 'ERROR': 'red'}.get(log['level'], 'gray')
                    st.markdown(f"<small><span style='color:{color}'>[{log['level']}]</span> {log['time']} - {log['message']}</small>", unsafe_allow_html=True)
                if st.button("🗑️ Clear Logs", key="clear_log"):
                    clear_logs()
                    st.rerun()
            else:
                st.caption("No logs yet")

    # Header - dynamic based on client (now after sidebar so we have dcs_name and pha_tool)
    st.markdown("""
    <div class="main-header">
        <h1>🔔 Alarm Rationalization Platform</h1>
        <p>Transform alarm management databases between DCS and PHA-Pro formats</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Main content
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 📁 Upload Files")
        
        if direction == "forward":
            # Determine file types based on client parser
            parser_type = client_config.get("parser", "dynamo")
            
            if parser_type == "abb":
                st.markdown(f"**Drag & drop** your {dcs_name} export file below, or click to browse")
                uploaded_file = st.file_uploader(
                    f"📂 {dcs_name} Export (.xlsx)",
                    type=['xlsx', 'xls'],
                    help=f"The Excel file exported from {dcs_name} containing alarm configuration",
                    key=f"forward_abb_{st.session_state.file_uploader_key}"
                )
                st.caption("Supported formats: .xlsx, .xls")
                # ABB doesn't need unit detection - it uses fixed unit
                unit_filter = ""
                unit_method_choice = "fixed"
            else:
                st.markdown(f"**Drag & drop** your {dcs_name} export file below, or click to browse")
                uploaded_file = st.file_uploader(
                    "📂 Alarm Database Export (.csv)",
                    type=['csv'],
                    help=f"The CSV file exported from {dcs_name} containing _DCSVariable, _DCS, _Parameter schemas",
                    key=f"forward_dynamo_{st.session_state.file_uploader_key}"
                )
                st.caption("Supported format: .csv (must contain _DCSVariable, _DCS, _Parameter schemas)")
            
            # Unit detection and selection (only for DynAMo parser)
            if parser_type != "abb":
                unit_filter = ""
                unit_method_choice = "tag_prefix"  # default
                
                if uploaded_file is not None:
                    # Scan file for available units
                    raw_bytes = uploaded_file.read()
                    uploaded_file.seek(0)  # Reset for later use
                    
                    file_content = None
                    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                        try:
                            file_content = raw_bytes.decode(enc)
                            break
                        except UnicodeDecodeError:
                            continue
                    
                    if file_content:
                        # Extract units using all methods
                        units_by_prefix, units_by_asset_parent, units_by_asset_child = scan_for_units(file_content, selected_client)
                        
                        # Show unit detection results
                        st.markdown("### 📊 Units Detected")
                        
                        # For DynAMo clients, show all methods and let user choose
                        if parser_type == "dynamo":
                            col_a, col_b, col_c = st.columns(3)
                            
                            with col_a:
                                st.markdown("**By Tag Prefix:**")
                                if units_by_prefix:
                                    st.code(", ".join(sorted(units_by_prefix, key=lambda x: (len(x), x))))
                                else:
                                    st.write("None found")
                            
                            with col_b:
                                st.markdown("**By Asset Path (Parent):**")
                                if units_by_asset_parent:
                                    # Sort and display
                                    sorted_parents = sorted(units_by_asset_parent)
                                    st.code(", ".join(sorted_parents[:10]) + ("..." if len(sorted_parents) > 10 else ""))
                                    if len(sorted_parents) > 10:
                                        with st.expander(f"Show all {len(sorted_parents)} units"):
                                            st.code(", ".join(sorted_parents))
                                else:
                                    st.write("None found")
                            
                            with col_c:
                                st.markdown("**By Asset Path (Child):**")
                                if units_by_asset_child:
                                    sorted_children = sorted(units_by_asset_child)
                                    st.code(", ".join(sorted_children[:10]) + ("..." if len(sorted_children) > 10 else ""))
                                    if len(sorted_children) > 10:
                                        with st.expander(f"Show all {len(sorted_children)} units"):
                                            st.code(", ".join(sorted_children))
                                else:
                                    st.write("None found")
                            
                            # Let user choose method
                            unit_method_choice = st.radio(
                                "Which unit extraction method should be used?",
                                options=["tag_prefix", "asset_parent", "asset_child"],
                                format_func=lambda x: {
                                    "tag_prefix": f"Tag Prefix ({len(units_by_prefix)} units) - e.g., '17' from '17TI5879'",
                                    "asset_parent": f"Asset Path - Parent ({len(units_by_asset_parent)} units) - e.g., '17_FLARE' (consolidated)",
                                    "asset_child": f"Asset Path - Child ({len(units_by_asset_child)} units) - e.g., '17H-2' (detailed)"
                                }[x],
                                help="""
**Tag Prefix**: Uses first 2 digits of tag name (e.g., 17TI5879 → 17)

**Asset Path - Parent**: Uses the first level after /U##/ in the asset hierarchy. 
This gives you consolidated units like 17_FLARE, 17_FGS, 17_ELEC.
Best for PHA-Pro import when you want fewer, larger unit groupings.

**Asset Path - Child**: Uses the last level in the asset hierarchy.
This gives you detailed units like 17H-2, 17IB-02, 17Z-50A.
Best when you need granular unit breakdown.
""",
                                horizontal=True
                            )
                            
                            # Show the units for selected method
                            if unit_method_choice == "tag_prefix":
                                available_units = units_by_prefix
                            elif unit_method_choice == "asset_parent":
                                available_units = units_by_asset_parent
                            else:
                                available_units = units_by_asset_child
                        # Note: ABB clients don't reach this code path (they use fixed units)
                        
                        st.markdown("---")
                
                # Unit filter input
                unit_filter = st.text_input(
                    "Filter by Unit(s)",
                    placeholder="e.g., 67 or 67,68,70 (leave blank for all)",
                    help="Enter unit numbers separated by commas to filter. Leave blank to process all units."
                )

                # Mode detection and selection (DynAMo clients only)
                if uploaded_file is not None and file_content:
                    detected_modes = scan_for_modes(file_content)

                    if detected_modes:
                        st.markdown("### 🔄 Mode Selection")

                        sorted_modes = sorted(detected_modes)
                        mode_count_str = ", ".join(sorted_modes)
                        st.caption(f"Modes found in file: {mode_count_str}")

                        # "All Modes" checkbox
                        all_modes = st.checkbox(
                            "Select All Modes",
                            value=True,
                            help="When checked, all modes are processed. Uncheck to select specific modes.",
                            key="all_modes_forward"
                        )

                        if all_modes:
                            selected_modes_ui = sorted_modes
                        else:
                            selected_modes_ui = st.multiselect(
                                "Select Mode(s) to Process",
                                options=sorted_modes,
                                default=["NORMAL"] if "NORMAL" in sorted_modes else sorted_modes[:1],
                                help="""**What are modes?**\n
Alarm databases can contain multiple configurations (modes) for the same tag/alarm.\n
- **NORMAL** — Active operating configuration (most common for rationalization)\n
- **Base** — Baseline template configuration\n
- **IMPORT / Export** — Administrative/transfer configurations\n
- **Startup / Shutdown** — Special operating state configurations\n\n
**When to use "All":** Select all modes if your file only contains one mode (e.g., only "Base") or if you need to process all configurations.\n\n
**When to select specific modes:** If your file has multiple modes and you only want to rationalize the active configuration, select just "NORMAL"."""
                            )

                        # Store in session state
                        st.session_state['selected_modes_forward'] = selected_modes_ui

                        # Show count
                        if selected_modes_ui:
                            st.info(f"ℹ️ Processing {len(selected_modes_ui)} mode(s): {', '.join(selected_modes_ui)}")
                        else:
                            st.warning("⚠️ No modes selected — transformation will produce no output.")
                    else:
                        st.session_state['selected_modes_forward'] = None
                
                # Store the method choice in session state for use during transform
                if 'unit_method_choice' not in st.session_state:
                    st.session_state.unit_method_choice = "tag_prefix"
                if uploaded_file is not None and parser_type == "dynamo":
                    st.session_state.unit_method_choice = unit_method_choice
            else:
                # ABB uses fixed unit from config - only show after file uploaded
                if uploaded_file is not None:
                    unit_value = client_config.get('unit_value', 'Line 1')
                    st.markdown(f"### 📊 Unit: **{unit_value}**")
                    st.markdown("---")
            
            source_file = None
            
        else:
            st.markdown(f"**Drag & drop** your {pha_tool} export file below, or click to browse")
            uploaded_file = st.file_uploader(
                f"📂 {pha_tool} MADB Export (.csv)",
                type=['csv'],
                help=f"The CSV file exported from {pha_tool} Alarm Management Database",
                key=f"reverse_phapro_{st.session_state.file_uploader_key}"
            )
            st.caption("Supported format: .csv (PHA-Pro MADB export)")

            # For DynAMo clients, require original file for mode preservation
            parser_type = client_config.get("parser", "dynamo")
            if parser_type == "dynamo":
                st.markdown("---")
                st.markdown(f"**⚠️ Required: Original {dcs_name} export file**")
                st.caption("Drag & drop or click to browse. Required to preserve client-specific values.")
                source_file = st.file_uploader(
                    f"📂 Original {dcs_name} Export (.csv)",
                    type=['csv'],
                    help=f"Upload the original {dcs_name} export to preserve client-specific configuration values.",
                    key=f"reverse_source_{st.session_state.file_uploader_key}"
                )

                if uploaded_file is not None and source_file is None:
                    st.warning(f"⚠️ Please upload the original {dcs_name} export file. Without it, default values will be used.")

                # Mode detection for reverse transform (from source file)
                if source_file is not None:
                    source_raw_peek = source_file.read()
                    source_file.seek(0)

                    source_content_peek = None
                    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                        try:
                            source_content_peek = source_raw_peek.decode(enc)
                            break
                        except UnicodeDecodeError:
                            continue

                    if source_content_peek:
                        detected_modes_rev = scan_for_modes(source_content_peek)

                        if detected_modes_rev:
                            st.markdown("### 🔄 Mode Selection")

                            sorted_modes_rev = sorted(detected_modes_rev)
                            mode_count_str_rev = ", ".join(sorted_modes_rev)
                            st.caption(f"Modes found in source file: {mode_count_str_rev}")

                            all_modes_rev = st.checkbox(
                                "Select All Modes",
                                value=True,
                                help="When checked, all modes from the source file are included in the output. Uncheck to select specific modes.",
                                key="all_modes_reverse"
                            )

                            if all_modes_rev:
                                selected_modes_rev_ui = sorted_modes_rev
                            else:
                                selected_modes_rev_ui = st.multiselect(
                                    "Select Mode(s) to Process",
                                    options=sorted_modes_rev,
                                    default=["NORMAL"] if "NORMAL" in sorted_modes_rev else sorted_modes_rev[:1],
                                    help="""**What are modes?**\n
Alarm databases can contain multiple configurations (modes) for the same tag/alarm.\n
Select which modes to include in the DCS return file.\n\n
**Important:** The mode value in each row is preserved exactly as-is from the original file — it is never modified. This selection only controls which rows are included in the output.""",
                                    key="mode_multiselect_reverse"
                                )

                            st.session_state['selected_modes_reverse'] = selected_modes_rev_ui

                            if selected_modes_rev_ui:
                                st.info(f"ℹ️ Processing {len(selected_modes_rev_ui)} mode(s): {', '.join(selected_modes_rev_ui)}")
                            else:
                                st.warning("⚠️ No modes selected — transformation will produce no output.")
                        else:
                            st.session_state['selected_modes_reverse'] = None
            else:
                # ABB clients - optional source file for Change Report
                st.markdown("---")
                st.markdown(f"**📊 Optional: Original {dcs_name} export for Change Report**")
                st.caption("Upload to generate a Change Report comparing original vs rationalized values.")
                source_file = st.file_uploader(
                    f"📂 Original {dcs_name} Export (.xlsx) - Optional",
                    type=['xlsx', 'xls'],
                    help=f"Upload the original {dcs_name} Excel export to enable Change Report generation.",
                    key=f"reverse_source_abb_{st.session_state.file_uploader_key}"
                )
            
            unit_filter = None
    
    with col2:
        st.markdown("### 📋 Output Format")
        
        parser_type = client_config.get("parser", "dynamo")
        
        if direction == "forward":
            if parser_type == "abb":
                st.markdown(f"""
                **{pha_tool} 23-Column Import**
                - Hierarchical format
                - Unit/Tag/Alarm structure
                - Ready for MADB import
                """)
            else:
                st.markdown(f"""
                **{pha_tool} 45-Column Import**
                - Hierarchical format
                - Unit/Tag/Alarm structure
                - Ready for MADB import
                """)
        else:
            if parser_type == "abb":
                st.markdown(f"""
                **{dcs_name} 8-Column Return**
                - Flat format
                - Direct {dcs_name} import
                - Consolidated notes
                """)
            else:
                st.markdown(f"""
                **{dcs_name} _Parameter 42-Column**
                - Flat format
                - Direct {dcs_name} import
                - Mode preservation supported
                """)
    
    st.markdown("---")
    
    # Column verification and Transform button
    if uploaded_file is not None:
        # Get the appropriate headers for this client/direction
        parser_type = client_config.get("parser", "dynamo")
        temp_transformer = AlarmTransformer(selected_client, selected_area)
        
        if direction == "forward":
            # Forward: Show output columns that will be generated
            st.markdown("### ✅ Verify PHA-Pro Column Format")
            st.info("**Before transforming**, verify that your PHA-Pro template expects these columns in this exact order.")
            
            expected_headers = temp_transformer.get_phapro_headers()
            
            with st.expander(f"📋 View Expected PHA-Pro Output Columns ({len(expected_headers)} columns)", expanded=False):
                # Display in 3 columns for better readability
                num_cols = 3
                cols = st.columns(num_cols)
                items_per_col = (len(expected_headers) + num_cols - 1) // num_cols
                
                for col_idx, col in enumerate(cols):
                    start_idx = col_idx * items_per_col
                    end_idx = min(start_idx + items_per_col, len(expected_headers))
                    
                    with col:
                        for i in range(start_idx, end_idx):
                            st.markdown(f"`{i+1:2d}` {expected_headers[i]}")
                
                st.markdown("---")
                st.markdown("**📋 Copy-friendly list:**")
                st.code(",".join(expected_headers), language=None)
            
            columns_confirmed = st.checkbox(
                "✓ I confirm these columns match my PHA-Pro import template",
                key="forward_columns_confirmed"
            )
            
        else:
            # Reverse: Show required input columns from PHA-Pro export
            st.markdown("### ✅ Verify PHA-Pro Export Columns")
            st.info("**Before transforming**, verify that your PHA-Pro export file contains these required columns.")
            
            # Get required columns for reverse transform
            required_cols = temp_transformer.get_required_columns_info()
            expected_headers = temp_transformer.get_phapro_headers()
            
            with st.expander(f"📋 View Required PHA-Pro Input Columns ({len(expected_headers)} columns)", expanded=False):
                # Display in 3 columns for better readability
                num_cols = 3
                cols = st.columns(num_cols)
                items_per_col = (len(expected_headers) + num_cols - 1) // num_cols
                
                for col_idx, col in enumerate(cols):
                    start_idx = col_idx * items_per_col
                    end_idx = min(start_idx + items_per_col, len(expected_headers))
                    
                    with col:
                        for i in range(start_idx, end_idx):
                            st.markdown(f"`{i+1:2d}` {expected_headers[i]}")
                
                st.markdown("---")
                st.markdown("**🔑 Key columns used for transformation:**")
                for col_name, purpose in required_cols.items():
                    st.markdown(f"- **{col_name}**: _{purpose}_")
                
                st.markdown("---")
                st.markdown("**📋 Copy-friendly list:**")
                st.code(",".join(expected_headers), language=None)
            
            columns_confirmed = st.checkbox(
                "✓ I confirm my PHA-Pro export has these columns",
                key="reverse_columns_confirmed"
            )
        
        st.markdown("")

        # Data Validation Preview (optional - off by default)
        preview_enabled = st.checkbox(
            "🔍 Preview data before transforming",
            value=False,
            help="Analyze the uploaded file to see what will be processed without performing the transformation"
        )

        if preview_enabled:
            with st.expander("📊 Data Preview & Validation", expanded=True):
                try:
                    file_content = uploaded_file.getvalue().decode('utf-8', errors='replace')
                    uploaded_file.seek(0)  # Reset file pointer

                    # Get selected modes for preview
                    preview_modes = None
                    if direction == "forward":
                        preview_modes = st.session_state.get('selected_modes_forward', None)
                    else:
                        preview_modes = st.session_state.get('selected_modes_reverse', None)

                    preview_stats = _preview_file_data(
                        file_content,
                        temp_transformer,
                        direction,
                        parser_type,
                        preview_modes
                    )

                    # Display preview stats
                    pcol1, pcol2, pcol3 = st.columns(3)

                    with pcol1:
                        st.metric("Total Rows", f"{preview_stats.get('total_rows', 0):,}")
                    with pcol2:
                        st.metric("Rows to Process", f"{preview_stats.get('rows_to_process', 0):,}")
                    with pcol3:
                        st.metric("Rows to Skip", f"{preview_stats.get('rows_to_skip', 0):,}")

                    # Show units found (for forward transform)
                    if direction == "forward" and preview_stats.get('units_found'):
                        st.markdown("**Units Found:**")
                        units_str = ", ".join(sorted(preview_stats['units_found']))
                        st.code(units_str)

                    # Show potential issues
                    if preview_stats.get('issues'):
                        st.warning("**Potential Issues Detected:**")
                        for issue in preview_stats['issues']:
                            st.markdown(f"- {issue}")
                    else:
                        st.success("No issues detected - data looks ready for transformation")

                    # Show skip reasons
                    if preview_stats.get('skip_reasons'):
                        with st.expander("View skipped row details"):
                            for reason, count in preview_stats['skip_reasons'].items():
                                st.markdown(f"- **{reason}**: {count:,} rows")

                except Exception as e:
                    st.error(f"Could not preview file: {str(e)}")

        st.markdown("")

        col1, col2, col3 = st.columns([1, 1, 1])

        with col2:
            # Only enable Transform if columns are confirmed
            if columns_confirmed:
                transform_clicked = st.button(
                    "🚀 Transform",
                    use_container_width=True,
                    type="primary"
                )
            else:
                st.button(
                    "🚀 Transform",
                    use_container_width=True,
                    type="primary",
                    disabled=True
                )
                st.caption("☝️ Please verify columns above before transforming")
                transform_clicked = False

        if transform_clicked:
            try:
                # Get parser type
                parser_type = client_config.get("parser", "dynamo")
                
                # Create transformer
                transformer = AlarmTransformer(selected_client, selected_area)
                
                # Initialize variables for change report
                file_content = None
                source_data = None
                
                if direction == "forward":
                    # Progress bar for forward transformation
                    progress_bar = st.progress(0, text="Initializing...")

                    if parser_type == "abb":
                        # ABB uses Excel, read as bytes
                        progress_bar.progress(20, text="Reading Excel file...")
                        raw_bytes = uploaded_file.read()

                        progress_bar.progress(50, text="Transforming to PHA-Pro format...")
                        output_csv, stats = transformer.transform_forward_abb(raw_bytes)
                        output_filename = f"{selected_client.upper()}_{pha_tool}_Import.csv"

                        progress_bar.progress(100, text="Complete!")
                    else:
                        # DynAMo uses CSV
                        progress_bar.progress(15, text="Reading CSV file...")
                        raw_bytes = uploaded_file.read()

                        progress_bar.progress(25, text="Detecting encoding...")
                        file_content = None
                        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                            try:
                                file_content = raw_bytes.decode(enc)
                                break
                            except UnicodeDecodeError:
                                continue

                        if file_content is None:
                            progress_bar.empty()
                            st.error("Could not decode file. Please ensure it's a valid CSV file.")
                            st.stop()

                        progress_bar.progress(40, text="Parsing alarm data...")

                        # Parse unit filter
                        selected_units = None
                        if unit_filter and unit_filter.strip():
                            selected_units = [u.strip() for u in unit_filter.split(',')]

                        # Get unit method from session state (for FLNG)
                        unit_method = st.session_state.get('unit_method_choice', 'tag_prefix')

                        progress_bar.progress(60, text="Transforming to PHA-Pro format...")

                        # Transform
                        # Get selected modes from session state
                        selected_modes = st.session_state.get('selected_modes_forward', None)
                        output_csv, stats = transformer.transform_forward(file_content, selected_units, unit_method, selected_modes)
                        output_filename = f"{selected_client.upper()}_{pha_tool}_Import.csv"

                        progress_bar.progress(100, text="Complete!")

                    # Clear progress bar after short delay
                    import time
                    time.sleep(0.3)
                    progress_bar.empty()
                    
                else:
                    # Reverse transformation - progress bar
                    progress_bar = st.progress(0, text="Initializing...")
                    progress_bar.progress(10, text="Reading PHA-Pro export...")
                    raw_bytes = uploaded_file.read()

                    progress_bar.progress(20, text="Detecting encoding...")
                    file_content = None
                    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                        try:
                            file_content = raw_bytes.decode(enc)
                            break
                        except UnicodeDecodeError:
                            continue

                    if file_content is None:
                        progress_bar.empty()
                        st.error("Could not decode file. Please ensure it's a valid CSV file.")
                        st.stop()

                    if parser_type == "abb":
                        # ABB reverse transformation
                        progress_bar.progress(50, text="Transforming ABB data...")
                        output_csv, stats = transformer.transform_reverse_abb(file_content)
                        output_filename = f"{selected_client.upper()}_{dcs_name}_Return.csv"
                        progress_bar.progress(100, text="Complete!")
                    else:
                        # DynAMo reverse transformation
                        progress_bar.progress(30, text="Reading original DynAMo file...")

                        # Load source data (full rows) from original file
                        source_data = None
                        if source_file:
                            source_raw = source_file.read()
                            source_content = None
                            for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                                try:
                                    source_content = source_raw.decode(enc)
                                    break
                                except UnicodeDecodeError:
                                    continue

                            if source_content:
                                progress_bar.progress(45, text="Parsing source alarm rows...")
                                # Parse ALL _Parameter rows from source file
                                source_rows = []
                                lines = source_content.replace('\r\n', '\n').split('\n')
                                reader = csv.reader(lines)
                                for row in reader:
                                    if len(row) >= 6 and row[0] == "_Variable" and row[2] == "_Parameter":
                                        source_rows.append(row)

                                source_data = {'rows': source_rows}


                        if not source_data:
                            progress_bar.empty()
                            st.error(f"Original {dcs_name} export file is required for reverse transformation.")
                            st.stop()

                        progress_bar.progress(60, text=f"Transforming {len(source_data['rows']):,} alarm rows...")

                        # Transform (merge PHA-Pro changes with original data)
                        # Get selected modes from session state
                        selected_modes_rev = st.session_state.get('selected_modes_reverse', None)
                        output_csv, stats = transformer.transform_reverse(file_content, source_data, selected_modes_rev)
                        output_filename = f"{selected_client.upper()}_{dcs_name}_Return.csv"

                        progress_bar.progress(100, text="Complete!")

                    # Clear progress bar after short delay
                    import time
                    time.sleep(0.3)
                    progress_bar.empty()
                # Show success (only after spinner completes)
                st.markdown("""
                <div class="status-success">
                    <strong>✅ Transformation Complete!</strong>
                </div>
                """, unsafe_allow_html=True)
                
                # Stats
                st.markdown("### 📊 Results")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown(f"""
                    <div class="stat-box">
                        <div class="stat-number">{stats['tags']:,}</div>
                        <div class="stat-label">Tags Processed</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"""
                    <div class="stat-box">
                        <div class="stat-number">{stats['alarms']:,}</div>
                        <div class="stat-label">Alarms Processed</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    # For reverse transform, show updated count; for forward, show units
                    if 'updated' in stats:
                        st.markdown(f"""
                        <div class="stat-box">
                            <div class="stat-number">{stats['updated']:,}</div>
                            <div class="stat-label">Alarms Updated</div>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        units_str = len(stats.get('units', set())) if isinstance(stats.get('units'), set) else "N/A"
                        st.markdown(f"""
                        <div class="stat-box">
                            <div class="stat-number">{units_str}</div>
                            <div class="stat-label">Units Found</div>
                        </div>
                        """, unsafe_allow_html=True)
                
                # Show not_found warning if any
                if stats.get('not_found', 0) > 0:
                    st.warning(f"⚠️ {stats['not_found']:,} alarms from original file were not found in PHA-Pro export (kept unchanged)")
                
                # Show skipped modes info with expandable explanation
                if stats.get('skipped_modes', 0) > 0:
                    skipped_mode_label = "non-selected modes" if st.session_state.get('selected_modes_forward') or st.session_state.get('selected_modes_reverse') else "non-NORMAL modes"
                    st.info(f"ℹ️ {stats['skipped_modes']:,} rows skipped ({skipped_mode_label})")
                    
                    with st.expander("📝 Click here to understand why rows were skipped"):
                        st.markdown("""
### What are "Modes" in DynAMo?

DynAMo uses **modes** to manage alarm configurations across different plant operating states. Each alarm can have different settings depending on which mode the plant is operating in:

| Mode | Purpose |
|------|---------|
| **NORMAL** | Standard operating conditions - this is the primary/active configuration |
| **IMPORT** | Used during data import operations |
| **Export** / **EXPORT** | Used during data export operations |
| **Base** | Baseline configuration template |
| **Startup** | Special settings during plant startup |
| **Shutdown** | Special settings during plant shutdown |

### Why are non-NORMAL rows skipped?

1. **The NORMAL mode is the active configuration** - When DynAMo is running in normal operations, it uses the NORMAL mode settings. The PHA-Pro rationalization process focuses on these active alarm configurations.

2. **Other modes are system/administrative rows** - Rows with modes like IMPORT, Export, EXPORT are typically:
   - Temporary configurations used during data transfers
   - Backup/snapshot configurations
   - Not actively used for alarm management

3. **Prevents duplicate alarms** - Your original DynAMo export file contains multiple rows for the same (tag, alarm type) combination - one for each mode. If we included all modes, the output would have duplicate entries that would cause import errors.

4. **Matches the manual process** - The manual Excel-based workflow also filters to only include NORMAL mode rows in the final import file.

### What this means for your output:

- ✅ **{:,} alarm rows with mode=NORMAL** were processed and included in the output
- â­ï¸ **{:,} rows with other modes** were skipped (they exist in your source file but are not part of the active alarm configuration)

The output file contains exactly one row per (tag, alarm type) combination, matching what DynAMo expects for a clean import.
                        """.format(stats['alarms'], stats['skipped_modes']))
                
                # P&ID Review Note (only for forward transformation)
                if direction == "forward":
                    st.warning("""
⚠️ **P&ID Review Required**

Before importing to PHA-Pro, please review and consolidate P&ID references:
- Tags without P&ID data are marked as "UNKNOWN"
- Verify P&ID assignments are correct for your facility
- Consolidate P&ID naming conventions if needed
""")
                
                # Download buttons
                st.markdown("### 📥 Download")

                col_dl1, col_dl2, col_dl3 = st.columns(3)

                with col_dl1:
                    st.download_button(
                        label=f"📄 CSV",
                        data=output_csv,
                        file_name=output_filename,
                        mime="text/csv",
                        use_container_width=True
                    )

                with col_dl2:
                    try:
                        excel_data = csv_to_excel(output_csv)
                        excel_filename = output_filename.replace('.csv', '.xlsx')
                        st.download_button(
                            label=f"📊 Excel",
                            data=excel_data,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    except Exception:
                        st.caption("Excel export unavailable")

                # Change Report button (for reverse transforms with source file)
                with col_dl3:
                    if direction == "reverse" and source_file:
                        try:
                            if parser_type == "abb":
                                # ABB Change Report - needs original Excel bytes
                                source_file.seek(0)
                                source_bytes = source_file.read()
                                change_report = transformer.generate_change_report_abb(file_content, source_bytes)
                            else:
                                # DynAMo Change Report
                                selected_modes_rep = st.session_state.get('selected_modes_reverse', None)
                                change_report = transformer.generate_change_report(file_content, source_data, selected_modes_rep)
                            report_filename = f"{selected_client.upper()}_{dcs_name}_Change_Report.xlsx"
                            st.download_button(
                                label="📋 Change Report",
                                data=change_report,
                                file_name=report_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        except Exception as report_error:
                            st.warning(f"Could not generate change report: {report_error}")
                
                # Preview
                with st.expander("👁️ Preview Output (first 20 rows)"):
                    # Handle both bytes and string output
                    if isinstance(output_csv, bytes):
                        preview_df = pd.read_csv(io.BytesIO(output_csv), nrows=20, encoding='latin-1')
                    else:
                        preview_df = pd.read_csv(io.StringIO(output_csv), nrows=20)
                    st.dataframe(preview_df, use_container_width=True)

                # Add to transformation history
                input_name = uploaded_file.name if uploaded_file else "unknown"
                output_bytes = output_csv if isinstance(output_csv, bytes) else output_csv.encode('latin-1', errors='replace')
                add_to_history(input_name, direction, selected_client, stats, output_bytes, output_filename)

            except Exception as e:
                    error_msg = str(e)
                    
                    # Check if this is a missing columns error
                    if error_msg.startswith("MISSING_COLUMNS:"):
                        missing_cols = error_msg.replace("MISSING_COLUMNS:", "").split(",")
                        
                        st.markdown("""
                        <div class="status-warning">
                            <strong>⚠️ Missing Required Columns</strong><br>
                            Your PHA-Pro export is missing columns needed for the DynAMo transformation.
                        </div>
                        """, unsafe_allow_html=True)
                        
                        st.markdown("### 📋 Missing Columns")
                        st.markdown("Please rename or add the following columns in your PHA-Pro export file:")
                        
                        # Get column descriptions
                        col_info = transformer.get_required_columns_info()
                        
                        # Create a nice table
                        missing_data = []
                        for col in missing_cols:
                            description = col_info.get(col, "Required for transformation")
                            missing_data.append({"Column Name": col, "Purpose": description})
                        
                        st.table(missing_data)
                        
                        st.markdown("---")
                        st.markdown("**How to fix:**")
                        st.markdown("""
                        1. Open your PHA-Pro export in Excel
                        2. Rename column headers to match the **exact names** shown above
                        3. Column names are case-sensitive and must match exactly
                        4. Save and re-upload the file
                        """)
                        
                        st.info("💡 **Note:** Column names must match exactly, including capitalization and special characters like parentheses.")
                    else:
                        st.error(f"Error during transformation: {error_msg}")
                        st.exception(e)
    
    else:
        st.markdown("""
        <div class="status-info">
            <strong>👆 Upload a file to get started</strong><br>
            Select your client profile and transformation direction in the sidebar, then upload your CSV file.
        </div>
        """, unsafe_allow_html=True)
    
    # Footer
    st.markdown("---")
    
    # Privacy Notice
    with st.expander("🔒 Privacy & Security"):
        st.markdown("""
        **Your data is secure.**
        
        **Access Control:**
        - ✅ **Authentication required** - No anonymous access to the tool
        - ✅ **Login protected** - Client names and functionality hidden until authenticated
        
        **Data Handling:**
        - ✅ **No data storage** - Files exist only in memory during your session
        - ✅ **No database** - Nothing is saved to any server
        - ✅ **No logging** - File contents are never logged
        - ✅ **Session isolation** - Your session is separate from other users
        - ✅ **Memory cleared** - All data erased when you log out or close the tab
        
        **In Transit:**
        - ✅ **HTTPS encrypted** - All traffic is encrypted
        
        **What is NOT exposed to the internet:**
        - Uploaded alarm data
        - Tag names or setpoints
        - Any client operational data
        - Client names (hidden behind login)
        
        Your alarm data, tag names, setpoints, and all operational information remain private and are never stored or exposed.
        
        *For additional security, this tool can be deployed on your own infrastructure behind your firewall.*
        """)
    
    st.markdown(
        "<div style='text-align: center; color: #6c757d; font-size: 0.85rem;'>"
        "Alarm Rationalization Platform • Applied Engineering Solutions • "
        f"Built with Streamlit • {datetime.now().year}"
        "</div>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()

